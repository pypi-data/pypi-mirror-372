"""
读取xlsx或xls文件

pip install xlrd
pip install xlwt
pip install openpyxl
"""

import openpyxl
import os
import xlrd
import xlwt
from collections.abc import Iterable
from collections.abc import Iterator
import datetime
import re


# 读取文件
class ReadExcelUtil(Iterable):

    def __init__(self, file_name, start_read_row_num=0):
        self.filename = file_name
        self.suffix = os.path.splitext(self.filename)[-1][1:].lower() # xls xlsx
        self.start_read_row_num = start_read_row_num # 开始从第N行读取数据
            
        
        if self.suffix == 'xlsx':
            self.workbook = openpyxl.load_workbook(self.filename, data_only=True)
            self.sheetnames = self.workbook.sheetnames # Sheet1 Sheet2 Sheet3...
            self.sheet = self.workbook[self.sheetnames[0]] # 根据名称默认读取第一个Sheet

        if self.suffix == 'xls':
            self.workbook = xlrd.open_workbook(self.filename, encoding_override="UTF-8")
            self.sheetnames = self.workbook._sheet_names # Sheet1 Sheet2 Sheet3...
            # self.sheet = self.workbook[0] # 根据下标默认读取第一个Sheet
            self.sheet = self.workbook[self.sheetnames[0]] # 根据名称默认读取第一个Sheet

    def __iter__(self):
        if self.suffix == 'xlsx':
            return ReadXlsxIterator(self.sheet, self.start_read_row_num)
        if self.suffix == 'xls':
            return ReadXlsIterator(self.sheet, self.start_read_row_num)


# 迭代器 xlsx
class ReadXlsxIterator(Iterator):

    def __init__(self, sheet, start_read_row_num):
        self.sheet = sheet
        self.max_row = sheet.max_row
        self.max_column = sheet.max_column
        self.row_num = start_read_row_num

    def __next__(self):
        try:
            if self.row_num <= self.max_row:
                # 读取行内容
                row_data = []
                for i in range(1, self.max_column + 1):
                    cell_value = self.sheet.cell(row=self.row_num, column=i).value
                    row_data.append(cell_value)
                
                rn = self.row_num
                self.row_num += 1
                # (行号, 行内容)
                return (rn, row_data)
        except Exception:
            raise StopIteration
        raise StopIteration


# 迭代器 xls
class ReadXlsIterator(Iterator):

    def __init__(self, sheet, start_read_row_num):
        self.sheet = sheet
        self.nrows = sheet.nrows
        self.row_num = start_read_row_num

    def __next__(self):
        try:
            if self.row_num <= self.nrows:
                # 读取行内容
                # (行号, 行内容)
                rn = self.row_num
                self.row_num += 1
                return (rn, self.sheet.row_values(rn))
        except Exception:
            raise StopIteration
        raise StopIteration


# 写入文件
class WriteXlsxUtil(object):

    def __init__(self, file_name, file_path='./'):
        self.index = 0
        self.suffix = '.xlsx'
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.create_sheet(index=self.index, title="Sheet")
        self.empty = True  # 标记是否真的写入了内容
        self.row = 0
        if file_name is None:
            self.file_name = "File-" + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + self.suffix
        else:
            self.file_name = file_name + self.suffix
        if file_path is None:
            self.file_path = './'
        else:
            self.file_path = file_path

    
    def set_sheet(self, sheet_name):
        self.index = self.index + 1
        self.sheet = self.workbook.create_sheet(index=self.index, title=sheet_name)
        self.row = 0
    
    # 设置列宽 index='A' width=30
    def column_width(self, index, width):
        self.sheet.column_dimensions[index].width = width    

    # 写入单元格内容
    def write(self, content):
        self.empty = False
        self.row += 1
        for index in range(1, len(content) + 1):  # index = column
            if isinstance(content[index - 1], str) or isinstance(content[index - 1], (int, float)):
                v = str(content[index - 1])
                digit = re.match(r'^\d+(\.\d+)?$', v, re.I)
                if digit:
                    self.sheet.cell(self.row, index).value = v
                else:
                    self.sheet.cell(self.row, index).value = v
            else:
                cell = content[index - 1]
                self.sheet.cell(self.row, index).value = cell['value']
                if 'color' in cell:
                    self.sheet.cell(self.row, index).font = openpyxl.styles.Font(color=cell['color'])

    # 保存文件
    def save(self):
        if not self.empty:
            self.workbook.save(self.file_path + self.file_name)



# class WriteXlsUtil(object):
#
#     def __init__(self, file_name='File'):
#         self.suffix = '.xls'
#         self.workbook = xlwt.Workbook()
#         self.sheet = self.workbook.add_sheet("Sheet")
#         self.file_name = file_name
#         self.empty = True
#         self.row = -1
#
#     # 写入单元格内容
#     def write(self, content):
#         self.empty = False
#         self.row += 1
#         for index in range(len(content)):
#             cnt = str(content[index])
#             if cnt.isdigit():
#                 self.sheet.write(self.row, index, cnt.zfill(len(cnt)))
#             else:
#                 self.sheet.write(self.row, index, content[index])
#
#     # 保存文件
#     def save(self):
#         if not self.empty:
#             self.workbook.save("./" + self.file_name + "-" + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + self.suffix)




__all__ = ['ReadExcelUtil', 'WriteXlsxUtil']
