"""Summarization functionality for gene enrichment analysis."""

import json
from typing import Dict, List
from openai import OpenAI
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict

class SummarizeAnalyzer:
    """Class for summarizing gene enrichment analysis results."""

    def __init__(self, api_key: str, model: str = "gpt-4.1-mini"):
        """Initialize the summarize analyzer.
        Args:
            api_key: OpenAI API key
            model: OpenAI model to use
        """
        self.client = OpenAI(api_key=api_key)
        self.summarize_model = model
    
    def _get_full_term(self,  name: str, source: str, results: Dict) -> Dict:
        """Get the full term from the results."""
        try:
            for term in results['results'][source]:
                if name.lower() in term['name'].lower():
                    return term
        except KeyError:
            print(f"Term {name} not found in {source} results")
            return None
    
    def _combine_results(self, enrichr_results: Dict, toppfun_results: Dict, gprofiler_results: Dict, terms_per_source: int) -> Dict:
        """Combine enrichment results from all tools."""
        # Ontologies
        combined_results = {}

        for tool in [enrichr_results, gprofiler_results, toppfun_results]:
            for source in tool:

                if source not in combined_results:
                    combined_results[source] = defaultdict(dict)

                for term in tool[source]:
                    if term in combined_results[source]:
                        # Prefers Enrichr's name if it exists because enrichr is first
                        tool[source][term]['name'] = combined_results[source][term]['name']
                        # Prefers GProfiler's ID if it exists (enrichr_results doesn't include IDs)
                        if 'id' in combined_results[source][term]:
                            tool[source][term]['id'] = combined_results[source][term]['id']
                        # Superset of genes
                        if 'genes' in combined_results[source][term] and 'genes' in tool[source][term]:
                            tool[source][term]['genes'] = list(set(tool[source][term]['genes'] + combined_results[source][term]['genes']))

                    combined_results[source][term].update(tool[source][term])

        # Sort the results by smallest p-value
        for source in combined_results:
            combined_results[source] = [term for term in combined_results[source].values()]
            combined_results[source] = sorted(combined_results[source], key=lambda x: min(x.get('gprofiler_p_value', 1), x.get('toppfun_p_value', 1), x.get('enrichr_p_value', 1)))[:terms_per_source]

        return combined_results
        

    def group_results_by_theme(self,
                               enrichr_results: Dict,
                               toppfun_results: Dict,
                               gprofiler_results: Dict,
                               literature_results: List[Dict],
                               gene_summaries: List[Dict],
                               terms_per_source: int,
                               genes: List[str],
                               ranked: bool,
                               context: str) -> str:
        """Group enrichment results into functional themes across all tools.
        
        Args:
            enrichr_results: Results from Enrichr analysis
            toppfun_results: Results from ToppFun analysis
            gprofiler_results: Results from gProfiler analysis
            literature_results: Results from PubMed analysis
            terms_per_source: Number of terms to retrieve per source
            genes: List of genes to analyze
            ranked: If the genes are ranked
            context: Additional user context
            
        Returns:
            Themed summary of results with each theme containing related terms
        """
        combined_results = self._combine_results(enrichr_results, toppfun_results, gprofiler_results, terms_per_source)
        combined_results['PubMed'] = literature_results
        if len(gene_summaries) > 0:
            combined_results['NCBI Gene Summaries'] = gene_summaries
        barcode_dict = dict()
        barcode = 100000
        for source in combined_results:
            for i, term in enumerate(combined_results[source]):
                term['barcode'] = barcode
                barcode_dict[barcode] = (source, i, term.pop('id'))
                barcode += 1

        # Pydantic model for output
        class LLMTheme(BaseModel):
            theme: str
            description: str
            confidence: float
            barcodes: List[int]

        class LLMThemedResults(BaseModel):
            themes: List[LLMTheme]
            summary: str

        # Send to LLM
        prompt = f"""You will be given a list of genes identified in a study{', ranked by differential expression.' if ranked else '.'}
        Your goal is to determine biological functions and pathways that may be consistently enriched in these genes, if any.
        You will also be given enrichment results from several different databases to help you with this task, including:
            Gene Ontology (GO), Human Phenotype (HP), KEGG, Reactome (REAC), WikiPathways (WP), Protein-Protein Interactions (PPI), Gene Summaries, and more.
        You will also be given a list of papers from PubMed that may be relevant to the genes.
        Each term will have a unique barcode, which you will use to identify the term in the enrichment results.
        
        Your task is to analyze the provided enrichment results and arrange them into consistent functional themes, if they exist.
        {'You should focus on themes that involve genes towards the top of the differential expression list.' if ranked else ''}
        Please delete any terms that don't coherently fit a theme, or that are not statistically significant.
        Strong results often have a p-value of 1E-10 or less, and a large number of genes (anything above half the number of genes in the list, or more than 10 genes).
        Results with a p-value of greater than 1E-5 or with a small number of genes (less than 10 or half the number of genes in the list) should be considered weak.
        For each theme, you should provide a confidence score between 0 and 1 (two decimal places), based on the strength of the evidence for the theme.
        If you're unsure about a theme, you should give it a confidence score of 0.5.
        If you're not confident about a theme, you should delete it.

        You will return a list of themes with the following attributes:
        * theme: The name of the theme
        * description: A brief description of the function of the theme and why you identified it (you don't need to include barcodes here)
        * barcodes: A list of integer barcodes, unique identifiers for the terms that are associated with the theme
        * confidence: A confidence score for the theme, between 0 and 1.
                      When determining confidence, you should consider the number of terms in the theme, the p-values of the terms, and the number of genes in the theme.
                      Strong results often have a p-value of 1E-10 or less, and a large number of genes.
                      Results with a p-value of greater than 1E-5 should be considered weak.
                      Be skeptical of themes that are not statistically significant, or that have a low number of genes.
                      Be skeptical of themes that are not consistent across the enrichment results.
        You will also provide a summary of the results, including a high level overview of what this gene list is enriched for.

        You should include literature terms in themes as they fit, but your final theme should be entitled "Literature Findings" and highlight
        interesting findings from PubMed, especially if they mention multiple genes from the list.
        If there are no coherent or consistent themes in the enrichment results, you should indicate that in the summary and return a single theme entitled "Literature Findings" that highlights interesting findings from PubMed.
        """
        
        response = self.client.responses.parse(
            model=self.summarize_model,
            input=[
                {"role": "system", "content": "You are an expert in molecular biology, immunology, oncology, and bioinformatics performing a functional enrichment analysis on a list of genes."},
                {"role": "user", "content": prompt},
                {"role": "user", "content": 'Context: ' + context},
                {"role": "user", "content": 'Genes: ' + ', '.join(genes)},
                {"role": "user", "content": 'Enrichment results:\n' + json.dumps(combined_results, indent=1)}
            ],
            text_format=LLMThemedResults
        )

        themed_results = response.output_parsed

        # Sort themes by confidence
        sorted_themes = sorted(themed_results.themes, key=lambda x: x.confidence, reverse=True)
        themed_results.themes = sorted_themes

        # Get full terms from barcodes
        clean_themed_results = {
            'summary': themed_results.summary,
            'themes': []
        }
        for theme in themed_results.themes:
            temp_theme = {
                'theme': theme.theme,
                'description': theme.description,
                'confidence': theme.confidence,
                'terms': []
            }
            for barcode in theme.barcodes:
                if barcode in barcode_dict:
                    source, i, term_id = barcode_dict[barcode]
                    term = combined_results[source][i]
                    term['id'] = term_id
                    term['source'] = source
                    temp_theme['terms'].append(term)
            clean_themed_results['themes'].append(temp_theme)

        return clean_themed_results

    def _sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """Sanitize a string to be used as an Excel sheet name.
        
        Excel sheet names cannot contain: : / \ ? * [ ]
        They also cannot start with a number or be longer than 31 characters.
        
        Args:
            name: The original name to sanitize
            max_length: Maximum length for the sheet name (default 31)
            
        Returns:
            Sanitized sheet name
        """
        # Replace invalid characters with spaces or underscores
        invalid_chars = [':', '/', '\\', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Remove leading/trailing whitespace
        sanitized = sanitized.strip()
        
        # If the name starts with a number, add a prefix
        if sanitized and sanitized[0].isdigit():
            sanitized = 'Theme_' + sanitized
        
        # Truncate to max length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]
        
        # Ensure it's not empty
        if not sanitized:
            sanitized = 'Theme'
            
        return sanitized

    def synthesize_analysis(self,
                            themed_results: Dict,
                            genes: List[str],
                            email: str,
                            search_terms: List[str],
                            context: str,
                            analysis_name: str,
                            results_dir: str) -> str:
        """Synthesize themed results into an Excel file with multiple sheets.
        
        Args:
            themed_results: Dict with the following keys:
                themes: List of themes, each containing:
                        - theme: Name of the theme
                        - description: Description of the theme
                        - terms: List of terms with tool-specific information
                summary: String with description of enrichment results
            genes: List of genes in analysis
            email: Email provided
            search_terms: List of search terms provided
            context: Context provided
            analysis_name: Name of the analysis
            results_dir: String with path to results
                
        Returns:
            Path to the generated Excel file
        """
        # Create a new workbook
        wb = Workbook()
        
        # Create summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Add metadata section
        summary_sheet['A1'] = "Analysis Metadata"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        summary_sheet['A3'] = "Email:"
        summary_sheet['B3'] = email
        summary_sheet['A3'].font = Font(bold=True)
        
        summary_sheet['A4'] = "Search Terms:"
        summary_sheet['B4'] = ", ".join(search_terms)
        summary_sheet['A4'].font = Font(bold=True)
        
        summary_sheet['A6'] = "Context:"
        summary_sheet['B6'] = context
        summary_sheet['A6'].font = Font(bold=True)
        summary_sheet['B6'].alignment = Alignment(wrap_text=True)
        
        summary_sheet['A8'] = "Genes:"
        summary_sheet['B8'] = ", ".join(genes)
        summary_sheet['A8'].font = Font(bold=True)
        summary_sheet['B8'].alignment = Alignment(wrap_text=True)
        
        # Add summary section
        summary_sheet['A10'] = "Analysis Summary"
        summary_sheet['A10'].font = Font(bold=True, size=14)
        summary_sheet['A11'] = "Description:"
        summary_sheet['A11'].font = Font(bold=True)
        summary_sheet['B11'] = themed_results['summary']
        summary_sheet['B11'].alignment = Alignment(wrap_text=True)
        summary_sheet['A13'] = "Themes"
        summary_sheet['A13'].font = Font(bold=True)
        summary_sheet['B13'] = "Name"
        summary_sheet['B13'].font = Font(bold=True)
        summary_sheet['C13'] = "Confidence"
        summary_sheet['C13'].font = Font(bold=True)
        current_row = 14

        found_lit = False
        for theme in themed_results['themes']:
            if 'Literature Findings' in theme['theme']:
                found_lit = True
                continue
            summary_sheet['B'+str(current_row)] = theme['theme']
            summary_sheet['C'+str(current_row)] = theme['confidence']
            # label confidence as high med or low
            if theme['confidence'] <= .87:
                confidence = "Low"
            elif theme['confidence'] <= .93:
                confidence = "Medium"
            else:
                confidence = "High"
            summary_sheet['D'+str(current_row)] = confidence
            summary_sheet['D'+str(current_row)].alignment = Alignment(horizontal='right')
            current_row += 1
        
        if found_lit:
            summary_sheet['B'+str(current_row)] = "Literature Findings"
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 100
        summary_sheet.column_dimensions['C'].width = 10
        
        # Create header row
        header = ["Name", "Source", "ID", "Genes", "Enrichr", "ToppFun", "gProfiler"]
        
        # Process each theme
        for theme in themed_results['themes']:
            if 'Literature Findings' in theme['theme']:
                continue
            # Create new sheet for theme
            theme_sheet = wb.create_sheet(title=self._sanitize_sheet_name(theme['theme']))
            
            # Add theme name
            theme_sheet['A1'] = theme['theme']
            theme_sheet['A1'].font = Font(bold=True)
            
            # Add theme description
            theme_sheet['A2'] = theme['description']
            theme_sheet['A2'].alignment = Alignment(wrap_text=True)

            # Add confidence
            theme_sheet['B2'] = f"Confidence: {theme['confidence']:.2f}"
            theme_sheet['B2'].font = Font(bold=True)
            
            # Add Terms header
            theme_sheet['A4'] = "Terms:"
            theme_sheet['A4'].font = Font(bold=True)
            
            # Add column headers
            for col, header_text in enumerate(header, 1):
                cell = theme_sheet.cell(row=5, column=col)
                cell.value = header_text
                cell.font = Font(bold=True)
            
            current_row = 6
            for term in theme['terms']:
                theme_sheet.cell(row=current_row, column=1, value=term['name'])
                theme_sheet.cell(row=current_row, column=2, value=term['source'])
                ID = int(term['id']) if term['source'] == 'PubMed' else term['id']
                theme_sheet.cell(row=current_row, column=3, value=ID)
                theme_sheet.cell(row=current_row, column=4, value=', '.join(term.get('genes', [])))
                theme_sheet.cell(row=current_row, column=5, value=term.get('enrichr_p_value', ' '))
                theme_sheet.cell(row=current_row, column=6, value=term.get('toppfun_p_value', ' '))
                theme_sheet.cell(row=current_row, column=7, value=term.get('gprofiler_p_value', ' '))
                current_row += 1
            
            # Adjust column widths
            theme_sheet.column_dimensions['A'].width = 80
            for col in range(2, len(header) + 1):
                theme_sheet.column_dimensions[chr(64 + col)].width = 20
        
        # Literature Findings theme
        if 'Literature Findings' in [theme['theme'] for theme in themed_results['themes']]:
            theme = next(theme for theme in themed_results['themes'] if theme['theme'] == 'Literature Findings')
            theme_sheet = wb.create_sheet(title=self._sanitize_sheet_name(theme['theme']))
                
            # Add theme name
            theme_sheet['A1'] = theme['theme']
            theme_sheet['A1'].font = Font(bold=True)
            
            # Add theme description
            theme_sheet['A2'] = theme['description']
            theme_sheet['A2'].alignment = Alignment(wrap_text=True)
            # Add headers for literature information
            headers = ['Paper Title', 'Year', 'PMID', 'Genes', 'Abstract']
            for col, header in enumerate(headers, 1):
                cell = theme_sheet.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True)
            
            # Add paper information
            current_row = 5
            for term in theme['terms']:
                if term['source'] == 'PubMed':
                    # Paper title
                    theme_sheet.cell(row=current_row, column=1, value=term['name'])
                    theme_sheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
                    
                    # Year
                    theme_sheet.cell(row=current_row, column=2, value=int(term.get('year')))
                    
                    # PMID
                    theme_sheet.cell(row=current_row, column=3, value=int(term.get('id')))

                    # Genes
                    genes = term.get('genes', [])
                    theme_sheet.cell(row=current_row, column=4, value=', '.join(genes))
                    
                    # Abstract
                    theme_sheet.cell(row=current_row, column=5, value=term.get('abstract', ''))
                    theme_sheet.cell(row=current_row, column=5).alignment = Alignment(wrap_text=True)
                    
                    current_row += 1
            
            # Adjust column widths
            theme_sheet.column_dimensions['A'].width = 80  # Title
            theme_sheet.column_dimensions['B'].width = 10  # Year
            theme_sheet.column_dimensions['C'].width = 15  # PMID
            theme_sheet.column_dimensions['D'].width = 20  # Genes
            theme_sheet.column_dimensions['E'].width = 100  # Abstract

        
        # Save the workbook
        output_path = f"{results_dir}/{analysis_name}_enrichment_analysis.xlsx"
        wb.save(output_path)
        return output_path 