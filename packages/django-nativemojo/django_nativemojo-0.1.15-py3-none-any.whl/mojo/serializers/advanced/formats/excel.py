try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAS_OPENPYXL = False

import io
from decimal import Decimal
from datetime import datetime, date, time
from django.http import HttpResponse
from django.db.models import QuerySet
from mojo.helpers import logit

logger = logit.get_logger("excel_formatter", "excel_formatter.log")


class ExcelFormatter:
    """
    Advanced Excel formatter with openpyxl integration and RestMeta.GRAPHS support.
    """
    
    def __init__(self, sheet_name="Sheet1", freeze_panes=True, auto_width=True, 
                 header_style=True, date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD HH:MM:SS"):
        """
        Initialize Excel formatter.
        
        :param sheet_name: Name for the Excel worksheet
        :param freeze_panes: Freeze header row for easier navigation
        :param auto_width: Auto-adjust column widths
        :param header_style: Apply styling to header row
        :param date_format: Format for date cells
        :param datetime_format: Format for datetime cells
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        self.sheet_name = sheet_name
        self.freeze_panes = freeze_panes
        self.auto_width = auto_width
        self.header_style = header_style
        self.date_format = date_format
        self.datetime_format = datetime_format
        
        # Style configurations
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
    
    def serialize_queryset(self, queryset, fields=None, graph=None, filename="export.xlsx", 
                          headers=None, localize=None):
        """
        Serialize a Django QuerySet to Excel format.
        
        :param queryset: Django QuerySet to serialize
        :param fields: List of field names or tuples (field_name, display_name)
        :param graph: RestMeta graph name to use for field configuration
        :param filename: Output filename
        :param headers: Custom header names (overrides field names)
        :param localize: Localization configuration
        :return: HttpResponse with Excel file
        """
        # Get fields configuration
        field_config = self._get_field_config(queryset, fields, graph)
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name
        
        # Write header row
        self._write_header_row(worksheet, field_config['headers'])
        
        # Write data rows
        row_num = 2  # Start after header
        for obj in queryset:
            try:
                row_data = self._extract_row_data(obj, field_config['field_names'], localize)
                self._write_data_row(worksheet, row_num, row_data)
                row_num += 1
            except Exception as e:
                logger.error(f"Error processing row for object {obj.pk}: {e}")
                continue
        
        # Apply formatting
        self._apply_formatting(worksheet, len(field_config['headers']), row_num - 1)
        
        # Create HTTP response
        return self._create_excel_response(workbook, filename)
    
    def serialize_data(self, data, fields=None, filename="export.xlsx", headers=None):
        """
        Serialize list of dictionaries or objects to Excel.
        
        :param data: List of dictionaries or objects
        :param fields: Field names to include
        :param filename: Output filename
        :param headers: Custom header names
        :return: HttpResponse with Excel file
        """
        if not data:
            return self._create_empty_excel_response(filename)
        
        # Auto-detect fields if not provided
        if not fields:
            fields = self._auto_detect_fields(data[0])
        
        # Prepare field configuration
        field_config = self._prepare_field_config(fields, headers)
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name
        
        # Write header row
        self._write_header_row(worksheet, field_config['headers'])
        
        # Write data rows
        for row_num, item in enumerate(data, start=2):
            try:
                row_data = self._extract_row_data(item, field_config['field_names'])
                self._write_data_row(worksheet, row_num, row_data)
            except Exception as e:
                logger.error(f"Error processing row {row_num}: {e}")
                continue
        
        # Apply formatting
        self._apply_formatting(worksheet, len(field_config['headers']), len(data) + 1)
        
        return self._create_excel_response(workbook, filename)
    
    def _get_field_config(self, queryset, fields, graph):
        """
        Get field configuration from various sources.
        """
        if fields:
            return self._prepare_field_config(fields)
        
        # Try to get from RestMeta.GRAPHS
        if graph and hasattr(queryset.model, 'RestMeta'):
            rest_meta = queryset.model.RestMeta
            if hasattr(rest_meta, 'GRAPHS') and graph in rest_meta.GRAPHS:
                graph_config = rest_meta.GRAPHS[graph]
                graph_fields = graph_config.get('fields', [])
                if graph_fields:
                    return self._prepare_field_config(graph_fields)
        
        # Fallback to model fields
        model_fields = [f.name for f in queryset.model._meta.fields]
        return self._prepare_field_config(model_fields)
    
    def _prepare_field_config(self, fields, headers=None):
        """
        Prepare field configuration for Excel generation.
        """
        field_names = []
        field_headers = []
        
        for field in fields:
            if isinstance(field, (tuple, list)):
                field_name, display_name = field
                field_names.append(field_name)
                field_headers.append(display_name)
            else:
                field_names.append(field)
                field_headers.append(field.replace('_', ' ').title())
        
        # Override with custom headers if provided
        if headers:
            field_headers = headers[:len(field_names)]
        
        return {
            'field_names': field_names,
            'headers': field_headers
        }
    
    def _write_header_row(self, worksheet, headers):
        """
        Write header row to worksheet with styling.
        """
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            
            if self.header_style:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
    
    def _write_data_row(self, worksheet, row_num, row_data):
        """
        Write data row to worksheet with proper type handling.
        """
        for col_num, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=row_num, column=col_num)
            
            # Handle different data types
            if isinstance(value, datetime):
                cell.value = value
                cell.number_format = self.datetime_format
            elif isinstance(value, date):
                cell.value = value
                cell.number_format = self.date_format
            elif isinstance(value, time):
                cell.value = value
                cell.number_format = "HH:MM:SS"
            elif isinstance(value, (int, float, Decimal)):
                try:
                    cell.value = float(value) if isinstance(value, Decimal) else value
                except (ValueError, TypeError):
                    cell.value = str(value)
            elif isinstance(value, bool):
                cell.value = value
            else:
                cell.value = str(value) if value is not None else ""
    
    def _apply_formatting(self, worksheet, num_cols, num_rows):
        """
        Apply formatting to the worksheet.
        """
        # Freeze panes (header row)
        if self.freeze_panes and num_rows > 1:
            worksheet.freeze_panes = 'A2'
        
        # Auto-adjust column widths
        if self.auto_width:
            self._auto_adjust_column_widths(worksheet, num_cols, num_rows)
    
    def _auto_adjust_column_widths(self, worksheet, num_cols, num_rows):
        """
        Auto-adjust column widths based on content.
        """
        for col_num in range(1, num_cols + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # Check all cells in the column
            for row_num in range(1, min(num_rows + 1, 100)):  # Limit to first 100 rows for performance
                try:
                    cell_value = worksheet[f"{column_letter}{row_num}"].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except Exception:
                    continue
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
    
    def _extract_row_data(self, obj, field_names, localize=None):
        """
        Extract row data from an object based on field names.
        """
        row = []
        
        for field_name in field_names:
            try:
                value = self._get_field_value(obj, field_name)
                value = self._process_field_value(value, field_name, localize)
                row.append(value)
            except Exception as e:
                logger.warning(f"Error extracting field '{field_name}': {e}")
                row.append(None)
        
        return row
    
    def _get_field_value(self, obj, field_name):
        """
        Get field value from object, supporting nested field access.
        """
        # Handle nested field access (e.g., "user.email", "profile.address.city")
        if '.' in field_name:
            return self._get_nested_field_value(obj, field_name)
        
        # Handle special metadata fields
        if field_name.startswith('metadata.') and hasattr(obj, 'getProperty'):
            parts = field_name.split('.', 2)
            if len(parts) == 3:
                return obj.getProperty(parts[2], category=parts[1])
            elif len(parts) == 2:
                return obj.getProperty(parts[1])
        
        # Standard field access
        if hasattr(obj, field_name):
            value = getattr(obj, field_name)
            return value() if callable(value) else value
        
        # Dictionary-style access
        if isinstance(obj, dict):
            return obj.get(field_name, None)
        
        return None
    
    def _get_nested_field_value(self, obj, field_path):
        """
        Get value from nested field path like "user.profile.name".
        """
        try:
            current = obj
            for field_part in field_path.split('.'):
                if current is None:
                    return None
                
                if hasattr(current, field_part):
                    current = getattr(current, field_part)
                elif isinstance(current, dict):
                    current = current.get(field_part)
                else:
                    return None
                
                # Handle callable attributes
                if callable(current):
                    current = current()
            
            return current
        except Exception as e:
            logger.warning(f"Error accessing nested field '{field_path}': {e}")
            return None
    
    def _process_field_value(self, value, field_name, localize=None):
        """
        Process field value with localization and special handling.
        """
        if value is None:
            return None
        
        # Apply localization if configured
        if localize and field_name in localize:
            try:
                localizer_config = localize[field_name]
                if '|' in localizer_config:
                    localizer_name, extra = localizer_config.split('|', 1)
                else:
                    localizer_name, extra = localizer_config, None
                
                # Import and apply localizer
                from mojo.serializers.formats.localizers import get_localizer
                localizer = get_localizer(localizer_name)
                if localizer:
                    return localizer(value, extra)
            except Exception as e:
                logger.warning(f"Localization failed for field '{field_name}': {e}")
        
        # Handle model instances
        if hasattr(value, 'pk'):
            return str(value)
        
        # Handle collections (flatten to string for Excel)
        elif isinstance(value, (list, tuple)):
            return '; '.join(str(item) for item in value)
        elif isinstance(value, dict):
            return str(value)
        
        return value
    
    def _auto_detect_fields(self, sample_item):
        """
        Auto-detect fields from a sample data item.
        """
        if isinstance(sample_item, dict):
            return list(sample_item.keys())
        elif hasattr(sample_item, '_meta'):
            return [f.name for f in sample_item._meta.fields]
        elif hasattr(sample_item, '__dict__'):
            return list(sample_item.__dict__.keys())
        else:
            return ['value']  # Fallback for primitive types
    
    def _create_excel_response(self, workbook, filename):
        """
        Create HTTP response with Excel file.
        """
        # Save workbook to BytesIO buffer
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = str(len(excel_buffer.getvalue()))
        
        return response
    
    def _create_empty_excel_response(self, filename):
        """
        Create response for empty dataset.
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name
        worksheet['A1'] = "No data available"
        
        return self._create_excel_response(workbook, filename)


# Convenience functions for backwards compatibility
def generate_excel(queryset, fields, filename, headers=None, localize=None, sheet_name="Sheet1"):
    """
    Generate Excel file from queryset.
    
    :param queryset: Django QuerySet
    :param fields: List of field names
    :param filename: Output filename
    :param headers: Custom header names
    :param localize: Localization config
    :param sheet_name: Excel sheet name
    :return: HttpResponse with Excel file
    """
    formatter = ExcelFormatter(sheet_name=sheet_name)
    return formatter.serialize_queryset(
        queryset=queryset,
        fields=fields,
        filename=filename,
        headers=headers,
        localize=localize
    )


def qsetToExcel(request, queryset, fields, filename, localize=None):
    """
    Legacy function name for backwards compatibility.
    """
    return generate_excel(queryset, fields, filename, localize=localize)


def serialize_to_excel(data, fields=None, filename="export.xlsx", headers=None, sheet_name="Sheet1"):
    """
    Serialize list of data to Excel.
    
    :param data: List of dictionaries or objects
    :param fields: Field names to include
    :param filename: Output filename
    :param headers: Custom header names
    :param sheet_name: Excel sheet name
    :return: HttpResponse with Excel file
    """
    formatter = ExcelFormatter(sheet_name=sheet_name)
    return formatter.serialize_data(
        data=data,
        fields=fields,
        filename=filename,
        headers=headers
    )


def create_multi_sheet_excel(data_sets, filename="export.xlsx"):
    """
    Create Excel file with multiple sheets.
    
    :param data_sets: List of tuples (sheet_name, queryset_or_data, fields, headers)
    :param filename: Output filename
    :return: HttpResponse with Excel file
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export")
    
    workbook = openpyxl.Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)
    
    for sheet_name, data, fields, headers in data_sets:
        formatter = ExcelFormatter(sheet_name=sheet_name)
        
        # Create new worksheet
        worksheet = workbook.create_sheet(title=sheet_name)
        
        if isinstance(data, QuerySet):
            field_config = formatter._get_field_config(data, fields, None)
        else:
            if not fields and data:
                fields = formatter._auto_detect_fields(data[0])
            field_config = formatter._prepare_field_config(fields, headers)
        
        # Write header
        formatter._write_header_row(worksheet, field_config['headers'])
        
        # Write data
        row_num = 2
        for item in data:
            try:
                row_data = formatter._extract_row_data(item, field_config['field_names'])
                formatter._write_data_row(worksheet, row_num, row_data)
                row_num += 1
            except Exception as e:
                logger.error(f"Error processing row in sheet '{sheet_name}': {e}")
                continue
        
        # Apply formatting
        formatter._apply_formatting(worksheet, len(field_config['headers']), row_num - 1)
    
    # Create response
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response