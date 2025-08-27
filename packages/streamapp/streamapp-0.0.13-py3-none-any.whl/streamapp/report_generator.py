""".xlsx files generator and .zip compresor for streamlit.

Create .xlsx files from scratch or with a template to
populate data and save them in a .zip bites object
to download from streamlit Download button.
"""

from io import BytesIO
from pathlib import Path
from datetime import date
from pandas import DataFrame
from typing import Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, NamedStyle, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from streamlit import secrets, toast
from zipfile import ZipFile, ZipInfo


class ReportGenerator:
    """.xlsx files generator

    Create .xlsx files from scratch or with a template to
    populate. New files are saved with tabular format,
    and the table style can be changed.

    This object use the stremalit secrets to define the
    templates folder with the variable name: `utils_files`

    For example if you save your template file in `.utils/consume`
    the varible is defined in the .toml file as:
    `utils_files = 'utils/consume'`
    """
    date = date.today().isoformat()
    # stylish
    totals_style = NamedStyle(name="totals_style")
    totals_style.font = Font(name='Arial', size=13, bold=True)
    totals_style.border = Border(
        top=Side(border_style='thick', color='FF000000'),
        bottom=Side(border_style=None, color='FF000000')
    )
    col = get_column_letter

    @staticmethod
    def __sheet_names(name: str):
        """Validate sheet names.

        Args:
            name: sheet or worksheet name

        Return:
            None
        """
        try:
            assert name.find(' ') == -1, 'Sheet names cannot have spaces'
        except AssertionError as e:
            return e
        return

    @staticmethod
    def __reports(workbook: Any, dfs: dict[str: DataFrame], file_name: str,
                  headers: bool = True, xslx_style: str = 'TableStyleMedium3',
                  initial_row: int = 1
                  ) -> list:
        """Generate or populate files with data.

        Args:
            workbook: An Openpyxl workbook object form a template
            dfs: a dict containing sheet name as key and Dataframe as
                data to be loaded.
            file_name: Download file`s name
            headers: Use the dataframe headers as firts row
            xslx_style: Excel table style to set in tables

        Return:
            list containing file name as first variable and bites type
            file object as second
        """

        for name, df in dfs.items():
            val = ReportGenerator.__sheet_names(name)
            if val:
                toast(val, icon='⛔')
                return [file_name+'.txt', b'']
            for i in df.columns:
                try:
                    df[i] = df[i].dt.date
                except Exception:
                    pass
            (max_row, max_col) = df.shape
            worksheet = workbook[name]

            for i in range(1, max_col+1):
                worksheet.column_dimensions[get_column_letter(i)].width = 23

            tab = Table(
                displayName=name,
                ref=f"A1:{get_column_letter(max_col)}{max_row+initial_row}"
            )
            tab.tableStyleInfo = TableStyleInfo(name=xslx_style)
            worksheet._tables.add(tab)

            base_df = dataframe_to_rows(df, index=False, header=headers)
            for c, r in enumerate(base_df, int(not headers)):
                for i, j in enumerate(r, 1):
                    worksheet.cell(c+1, i, j)

        with BytesIO() as buffer:
            workbook.save(buffer)
            file = buffer.getvalue()
        workbook.close()

        return [file_name+'.xlsx', file]

    @staticmethod
    def from_template_xlsx(file_name: str, dfs: dict[str: DataFrame],
                           base_file: str, sub_folder: str = '',
                           xslx_style: str = 'TableStyleMedium3',
                           initial_row: int = 1) -> list:
        """Populate and existing .xlsx template

        Args:
            dfs: a dict containing sheet name as key and Dataframe as
                data to be loaded.
            file_name: Download file`s name
            sub_folder: folder where the template is stored
            base_file: Template's name to populate without extension

        Returns:
            list containing file name as first variable and bites type
            file object as second
        """
        try:
            workbook = load_workbook(
                Path(
                    secrets.get('utils_files', ''),
                    sub_folder, base_file + '.xlsx'
                ).as_posix()
            )
            result = ReportGenerator.__reports(
                workbook=workbook,
                dfs=dfs,
                file_name=file_name,
                headers=False,
                xslx_style=xslx_style,
                initial_row=initial_row
            )
        except Exception:
            result = ReportGenerator.xlsx(
                dfs=dfs,
                file_name=file_name,
                xslx_style=xslx_style
            )

        return result

    @staticmethod
    def xlsx(dfs: dict[str, DataFrame], file_name: str,
             xslx_style: str = 'TableStyleMedium3') -> list:
        """Create a new .xlsx file.

        Args:
            dfs: a dict containing sheet name as key and Dataframe as
                data to be loaded.
            file_name: Download file`s name

        Returns:
            list containing file name as first variable and bites type
            file object as second
        """
        workbook = Workbook()
        for n, i in enumerate(dfs.keys()):
            val = ReportGenerator.__sheet_names(i)
            if val:
                toast(val, icon='⛔')
                return [file_name+'.txt', b'']
            if n == 0:
                workbook.active.__setattr__('title', i)
            else:
                workbook.create_sheet(title=i)
        result = ReportGenerator.__reports(
            workbook=workbook,
            dfs=dfs,
            file_name=file_name,
            xslx_style=xslx_style
        )

        return result


class InMemoryZip(object):
    """Create a .zip bites object to download from Download button
    in streamlit.

    It receives multiple files objects as bites
    """
    zip_file = BytesIO()

    @classmethod
    def create_zip(cls, files: list[list[str: bytes]]) -> bytes:
        """Generate a .zip object as bites with multiple
        files objects represented as bites too.

        Args:
            files: a list object containing list with the file name
                as firts variable an its bytes representation as second
        Returns:
            .zip folder represented as bites
        """
        with ZipFile(cls.zip_file, 'w') as zip_archive:
            for i in files:
                zip_archive.writestr(
                    zinfo_or_arcname=ZipInfo(i[0]),
                    data=i[1]
                )
        return cls.zip_file
