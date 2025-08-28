import tkinter as tk
import io
import pandas as pd
import re

import os
from shutil import copy2
import openpyxl
import traceback
import tkinter as tk  # 添加tkinter导入

# 重定向标准输出到一个字符串缓冲区
class StdoutRedirector(io.StringIO):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget

    def write(self, s):
        self.text_widget.insert(tk.END, s)
        self.text_widget.see(tk.END)  # 自动滚动到底部


# 根据文件类型提取数据
def extract_data(file_path, console_text):
    # 将文件扩展名转换为小写，以忽略大小写
    file_path = file_path.lower()
    if file_path.endswith(('.htm', '.html')):
        return extract_data_from_html(file_path)
    elif file_path.endswith('.txt') or file_path.endswith('.net'):
        return extract_data_from_file(file_path)
    elif file_path.endswith('.xlsx'):
        return read_excel(file_path)
    else:
        console_text.insert(tk.END, "不支持的文件类型。\n")
        return None

# 提取HTML表格数据
def extract_data_from_html(file_path):
    data = {}
    try:
        dfs = pd.read_html(file_path)
        if len(dfs) > 0:
            df = dfs[0]
            df = df.iloc[1:]  # 跳过第一行
            for _, row in df.iterrows():
                netname = row[0]
                netpin = row[1]
                if netname not in data:
                    data[netname] = []
                data[netname].append(netpin)
        return data
    except Exception as e:
        print(f"读取 HTML 文件时发生错误: {e}")
        return None

# 提取文本文件或.Net文件数据
def extract_data_from_file(file_path):
    data = {}
    with open(file_path, 'r') as file:
        lines = file.readlines()
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line == '(':
                i += 1
                if i < len(lines):
                    netname = lines[i].strip()
                    netpins = []
                    i += 1
                    while i < len(lines) and not lines[i].strip().startswith(')'):
                        netpin = lines[i].strip()
                        if netpin:
                            netpins.append(netpin)
                        i += 1
                    if i < len(lines) and lines[i].strip().startswith(')'):
                        i += 1
                    data[netname] = netpins
            else:
                i += 1
    return data

# 提取Excel文件数据
def read_excel(file_path):
    df = pd.read_excel(file_path)
    data = {}
    for _, row in df.iterrows():
        netname = row['Net Name']
        netpin = row['Net Pins']
        if netname not in data:
            data[netname] = []
        data[netname].append(netpin)
    return data

# 数据处理函数
def process_data(data):
    data_list = []
    for netname, netpins in data.items():
        netpins_str = ' '.join([''.join(pin) for pin in netpins])
        data_list.append([netname, netpins_str])
    first_pins = [(row[0], row[1].split()[0]) for row in data_list]
    sorted_first_pins = sorted(first_pins, key=lambda x: natural_keys(x[1]))
    sorted_data = []
    for net_name, first_pin in sorted_first_pins:
        for row in data_list:
            if row[0] == net_name and row[1].startswith(first_pin):
                sorted_data.append(row)
                break
    return sorted_data

# 保存到Excel
def save_to_excel(data, output_file):
    new_df = pd.DataFrame(data, columns=['Net Name', 'Net Pins'])
    new_df.to_excel(output_file, index=False)

# 模式四专用保存到Excel函数
def save_to_excel_mode_four(data, output_file):
    new_df = pd.DataFrame(data, columns=['Net Name', 'Net Pin1', 'Net Pin2', '导通电阻'])
    new_df.to_excel(output_file, index=False)

# region 模式五专用函数
# 模式五专用保存到Excel函数 - 连接模块分析
def save_to_excel_mode_five(original_data, connection_results, output_file):
    """保存模式五的连接模块分析结果到Excel"""
    import openpyxl
    from openpyxl.styles import PatternFill
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "连接模块分析"
    
    # 写入表头
    headers = ['Net Name', 'Net Pin1', 'Net Pin2', '导通电阻', '结果']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入原始数据
    for row_idx, data_row in enumerate(original_data, 2):
        for col_idx, value in enumerate(data_row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 定义颜色列表（用于连接模块背景色）
    colors = [
        'FFE6E6',  # 浅红色
        'E6F3FF',  # 浅蓝色  
        'E6FFE6',  # 浅绿色
        'FFF0E6',  # 浅橙色
        'F3E6FF',  # 浅紫色
        'FFFFE6',  # 浅黄色
        'E6FFFF',  # 浅青色
        'FFE6F3',  # 浅粉色
    ]
    
    # 应用连接模块的背景色 - 只对连接的两个单元格应用颜色
    color_index = 0
    last_used_color = -1
    
    for result in connection_results:
        connection_modules = result.get('connection_modules', [])
        
        for module in connection_modules:
            # 确保相邻连接模块颜色不同
            while color_index == last_used_color:
                color_index = (color_index + 1) % len(colors)
            
            fill = PatternFill(start_color=colors[color_index], 
                              end_color=colors[color_index], 
                              fill_type='solid')
            
            start_row = module['start_row'] + 2  # +2 因为Excel从1开始且有表头
            end_row = module['end_row'] + 2
            pin_type = module['pin_type']
            
            if pin_type == 'NetPin1':
                # NetPin1连接：对两行的第2列（Net Pin1）应用颜色
                ws.cell(row=start_row, column=2).fill = fill  # 起始行的Net Pin1
                ws.cell(row=end_row, column=2).fill = fill    # 结束行的Net Pin1
            else:  # NetPin2
                # NetPin2连接：对两行的第3列（Net Pin2）应用颜色
                ws.cell(row=start_row, column=3).fill = fill  # 起始行的Net Pin2
                ws.cell(row=end_row, column=3).fill = fill    # 结束行的Net Pin2
            
            last_used_color = color_index
            color_index = (color_index + 1) % len(colors)
    
    # 处理结果列 - 按分组分列（按前缀排序）
    import re
    
    # 首先收集所有的group_key并排序
    all_group_keys = set()
    for result in connection_results:
        all_group_keys.add(result['group_key'])
    
    # 对group_key进行自然排序
    def natural_sort_key(group_key):
        # 提取数字和字母部分
        parts = re.findall(r'(\d+|\D+)', group_key)
        result = []
        for part in parts:
            if part.isdigit():
                result.append(int(part))
            else:
                result.append(part)
        return result
    
    sorted_group_keys = sorted(all_group_keys, key=natural_sort_key)
    
    # 按排序后的顺序分配列
    group_columns = {}  # 存储每个分组对应的列号
    current_result_col = 5  # 结果列从第5列开始
    
    for group_key in sorted_group_keys:
        group_columns[group_key] = current_result_col
        # 写入分组标题
        ws.cell(row=1, column=current_result_col, value=f'结果-{group_key}')
        current_result_col += 1
    
    # 填入结果数据
    for result in connection_results:
        group_key = result['group_key']
        start_pin = result['start_pin']
        end_pin = result['end_pin']
        
        # 在对应的分组列中写入结果
        target_col = group_columns[group_key]
        
        # 找到该列的下一个空位
        target_row = 2
        while ws.cell(row=target_row, column=target_col).value is not None:
            target_row += 1
        
        # 写入头尾引脚对
        pin_pair = f"{start_pin} {end_pin}"
        ws.cell(row=target_row, column=target_col, value=pin_pair)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存工作簿
    wb.save(output_file)

# 读取包含导通电阻的Excel文件（模式五专用）
def read_excel_with_resistance(file_path, console_text):
    """读取包含Net Name、Net Pin1、Net Pin2、导通电阻列的Excel文件"""
    try:
        console_text.insert(tk.END, f"正在读取Excel文件: {file_path}\n")
        
        # 读取Excel文件
        df = pd.read_excel(file_path)
        console_text.insert(tk.END, f"Excel文件读取成功，数据行数: {len(df)}\n")
        
        # 检查必要的列是否存在
        required_columns = ['Net Name', 'Net Pin1', 'Net Pin2', '导通电阻']
        missing_columns = []
        
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        if missing_columns:
            console_text.insert(tk.END, f"Excel文件缺少必要的列: {', '.join(missing_columns)}\n")
            console_text.insert(tk.END, f"当前列: {', '.join(df.columns.tolist())}\n")
            return None
        
        # 处理数据
        result_data = []
        for index, row in df.iterrows():
            try:
                net_name = str(row['Net Name']).strip()
                net_pin1 = str(row['Net Pin1']).strip()
                net_pin2 = str(row['Net Pin2']).strip()
                
                # 处理导通电阻列，可能是数字或字符串
                resistance_value = row['导通电阻']
                if pd.isna(resistance_value) or str(resistance_value).strip() == '':
                    resistance = 0.0
                else:
                    try:
                        resistance = float(resistance_value)
                    except (ValueError, TypeError):
                        console_text.insert(tk.END, f"第{index+2}行导通电阻值无效: {resistance_value}，使用0.0\n")
                        resistance = 0.0
                
                # 跳过空行或无效行
                if not net_name or not net_pin1 or not net_pin2:
                    continue
                
                result_data.append([net_name, net_pin1, net_pin2, resistance])
                
            except Exception as e:
                console_text.insert(tk.END, f"处理第{index+2}行时出错: {str(e)}\n")
                continue
        
        console_text.insert(tk.END, f"成功解析 {len(result_data)} 行有效数据\n")
        
        # 显示前几行数据供确认
        if result_data:
            console_text.insert(tk.END, "前3行数据预览:\n")
            for i, row in enumerate(result_data[:3]):
                console_text.insert(tk.END, f"  {i+1}: {row[0]} | {row[1]} | {row[2]} | {row[3]}\n")
        
        return result_data
        
    except Exception as e:
        console_text.insert(tk.END, f"读取Excel文件出错: {str(e)}\n")
        return None


def analyze_connection_modules(data, k_value, console_text):
    """分析连接模块的核心函数"""
    results = []
    current_row = 0

    while current_row < len(data):
        console_text.insert(tk.END, f"\n开始新的连接分析，起始行: {current_row + 1}\n")

        # 从当前行开始进行连接分析
        connection_modules, end_row = find_connection_path(data, current_row, k_value, console_text)

        if connection_modules:
            # 获取头尾引脚
            start_row = connection_modules[0]['start_row']
            end_row = connection_modules[-1]['end_row']
            start_pin, end_pin = get_head_tail_pins_from_modules(data, connection_modules)

            # 根据引脚前缀分组
            group_key = get_pin_group_key(start_pin)

            # 计算总电阻：从起始行到结束行的累加
            total_resistance = sum([data[i][3] for i in range(start_row, end_row + 1)])

            result = {
                'start_row': start_row,
                'end_row': end_row,
                'start_pin': start_pin,
                'end_pin': end_pin,
                'group_key': group_key,
                'connection_modules': connection_modules,
                'total_resistance': total_resistance
            }
            results.append(result)

            console_text.insert(tk.END,
                                f"连接完成: {start_pin} - {end_pin}, 总电阻: {result['total_resistance']:.2f}\n")
            console_text.insert(tk.END,
                                f"连接模块数量: {len(connection_modules)}, 最后连接类型: {connection_modules[-1]['pin_type']}\n")

        # 移动到下一个未处理的行
        current_row = end_row + 1 if connection_modules else current_row + 1

    return results


def find_connection_path(data, start_row, k_value, console_text):
    """找到从start_row开始的连接路径"""
    console_text.insert(tk.END, f"\n=== 第一步：电阻累加分析 ===\n")

    # 第一步：累加电阻，找到不超过k值的最大行数范围
    cumulative_resistance = 0.0
    max_row = start_row

    for i in range(start_row, len(data)):
        cumulative_resistance += data[i][3]
        console_text.insert(tk.END, f"累加到第{i + 1}行，累计电阻: {cumulative_resistance:.2f}\n")

        if cumulative_resistance <= k_value:
            max_row = i
        else:
            console_text.insert(tk.END, f"超过阈值 {k_value}，回退到第{max_row + 1}行\n")
            break

    if max_row == start_row:
        console_text.insert(tk.END, "起始行就超过阈值，无法形成连接\n")
        return None, start_row

    # 第二步：确保参与行数为偶数
    participate_rows = max_row - start_row + 1
    console_text.insert(tk.END, f"\n=== 第二步：行数调整 ===\n")
    console_text.insert(tk.END, f"参与行数: {participate_rows} (第{start_row + 1}行到第{max_row + 1}行)\n")

    if participate_rows % 2 == 1:  # 奇数行
        max_row -= 1  # 后退一行
        participate_rows -= 1
        final_resistance = sum([data[i][3] for i in range(start_row, max_row + 1)])
        console_text.insert(tk.END, f"行数为奇数，后退一行到第{max_row + 1}行，参与行数: {participate_rows}\n")
        console_text.insert(tk.END, f"调整后累计电阻: {final_resistance:.2f}\n")
    else:
        final_resistance = sum([data[i][3] for i in range(start_row, max_row + 1)])
        console_text.insert(tk.END, f"行数为偶数，无需调整，累计电阻: {final_resistance:.2f}\n")

    # 第三步：生成连接模块
    console_text.insert(tk.END, f"\n=== 第三步：生成连接模块 ===\n")
    connection_modules = []
    connection_type = 1  # 1: NetPin1连接, 2: NetPin2连接

    for i in range(start_row, max_row):  # max_row不包含，因为连接是相邻行
        pin_type = 'NetPin1' if connection_type % 2 == 1 else 'NetPin2'
        console_text.insert(tk.END, f"连接模块{connection_type}: 第{i + 1}行{pin_type} - 第{i + 2}行{pin_type}\n")

        connection_modules.append({
            'type': connection_type,
            'start_row': i,
            'end_row': i + 1,
            'pin_type': pin_type,
        })
        connection_type += 1

    # 第四步：检查最后连接模块，如果不是NetPin1，后退2行
    console_text.insert(tk.END, f"\n=== 第四步：NetPin1结束检查 ===\n")
    if connection_modules:
        last_module = connection_modules[-1]
        console_text.insert(tk.END, f"最后连接模块类型: {last_module['pin_type']}\n")

        if last_module['pin_type'] != 'NetPin1':
            console_text.insert(tk.END, f"最后连接不是NetPin1，后退2行\n")
            # 移除最后2个连接模块
            if len(connection_modules) >= 2:
                removed1 = connection_modules.pop()
                removed2 = connection_modules.pop()
                console_text.insert(tk.END, f"移除连接模块{removed2['type']}和{removed1['type']}\n")

                # 重新计算最终电阻
                if connection_modules:
                    final_end_row = connection_modules[-1]['end_row']
                    final_resistance = sum([data[i][3] for i in range(start_row, final_end_row + 1)])
                else:
                    final_resistance = data[start_row][3]
            else:
                # 如果连接模块少于2个，清空
                console_text.insert(tk.END, f"连接模块不足2个，清空所有连接\n")
                connection_modules = []
                final_resistance = 0

        if connection_modules:
            console_text.insert(tk.END,
                                f"最终保留 {len(connection_modules)} 个连接模块，最后连接类型: {connection_modules[-1]['pin_type']}\n")
            console_text.insert(tk.END, f"最终累计电阻: {final_resistance:.2f}\n")

            end_row_result = connection_modules[-1]['end_row']
            return connection_modules, end_row_result

    console_text.insert(tk.END, "无有效连接模块\n")
    return None, start_row


def get_head_tail_pins_from_modules(data, connection_modules):
    """从连接模块信息获取头尾引脚"""
    if not connection_modules:
        return None, None

    # 第一个连接模块的起始行
    start_row = connection_modules[0]['start_row']
    # 最后一个连接模块的结束行
    end_row = connection_modules[-1]['end_row']

    # 起始引脚：第一个连接一定是NetPin1，所以未连接的是NetPin2
    start_pin = data[start_row][2]  # NetPin2

    # 结束引脚：最后一个连接一定是NetPin1（因为我们强制以NetPin1结束），所以未连接的是NetPin2
    end_pin = data[end_row][2]  # NetPin2

    return start_pin, end_pin


def get_head_tail_pins(data, connection_path):
    """获取连接路径的头尾引脚（兼容旧版本）"""
    if not connection_path:
        return None, None

    start_row = connection_path[0]
    end_row = connection_path[-1]

    # 连接数量
    connection_count = len(connection_path) - 1

    # 起始引脚：第一个连接是NetPin1，所以未连接的是NetPin2
    start_pin = data[start_row][2]  # NetPin2

    # 结束引脚：根据连接数量的奇偶性确定
    if connection_count % 2 == 1:  # 奇数个连接，最后连接的是NetPin1，未连接的是NetPin2
        end_pin = data[end_row][2]  # NetPin2
    else:  # 偶数个连接，最后连接的是NetPin2，未连接的是NetPin1
        end_pin = data[end_row][1]  # NetPin1

    return start_pin, end_pin


def get_pin_group_key(pin_name):
    """获取引脚的分组键（.前的数字字母部分）"""
    if '.' in pin_name:
        return pin_name.split('.')[0]
    return pin_name
# endregion

# 排序辅助函数
def natural_keys(text):
    parts = re.split(r'(\d+)', text)
    return [int(part) if part.isdigit() else part for part in parts]


# region xs前缀函数代码
# 为接点号添加前缀的辅助函数
def process_terminal_with_prefix(cell_value, prefix):
    """
    为接点号添加前缀
    """
    # 不同分隔符的处理
    separators = ['、', ' ', ',', '，', '.']

    # 清理并转换为字符串
    value = str(cell_value).strip()

    # 如果已经有前缀，不处理
    if re.search(fr'^{prefix}\.', value):
        return value

    # 检查是否包含分隔符
    for sep in separators:
        if sep in value:
            # 分割并处理每个部分
            parts = value.split(sep)
            processed_parts = []

            for part in parts:
                part = part.strip()
                # 只处理包含数字或字母的部分
                if part and (re.search(r'\d', part) or re.search(r'[a-zA-Z]', part)):
                    processed_parts.append(f"{prefix}.{part}")
                else:
                    processed_parts.append(part)

            return sep.join(processed_parts)

    # 检查单个值是否包含数字或字母
    if value and (re.search(r'\d', value) or re.search(r'[a-zA-Z]', value)):
        return f"{prefix}.{value}"

    # 其他情况原样返回
    return value


# 模式三的Excel处理函数
def mode_three_excel_process(input_file, output_file, console_text):
    """
    Excel处理函数：读取Excel文件，添加前缀并保留格式
    """
    try:
        console_text.insert(tk.END, f"正在处理Excel文件: {input_file}\n")

        # 检查文件格式
        file_ext = os.path.splitext(input_file)[1].lower()
        if file_ext not in ['.xlsx', '.xls']:
            console_text.insert(tk.END, "请选择Excel文件(*.xlsx, *.xls)\n")
            return

        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 确保输出文件扩展名与输入文件一致
        output_ext = os.path.splitext(output_file)[1].lower()
        if output_ext != file_ext:
            output_file = os.path.splitext(output_file)[0] + file_ext
            console_text.insert(tk.END, f"已调整输出文件扩展名为与输入一致: {output_file}\n")

        # 先复制原文件到目标位置
        try:
            copy2(input_file, output_file)
            console_text.insert(tk.END, f"已复制原始文件到: {output_file}\n")
        except Exception as e:
            console_text.insert(tk.END, f"复制文件失败: {str(e)}\n")
            return

        # 使用pandas读取Excel获取数据
        try:
            # 检查文件中是否包含Sheet1
            excel_file = pd.ExcelFile(input_file)
            sheet_names = excel_file.sheet_names
            console_text.insert(tk.END, f"Excel文件包含工作表: {sheet_names}\n")

            if 'Sheet1' not in sheet_names:
                console_text.insert(tk.END, "错误：Excel文件中不包含Sheet1工作表！\n")
                return

            # 读取Sheet1
            df = pd.read_excel(input_file, sheet_name='Sheet1', header=None)
            console_text.insert(tk.END, f"Sheet1工作表读取成功，形状: {df.shape}\n")

            # 查找连接点和接点号位置
            terminal_columns = []

            # 先查找连接点相关行
            connect_point_rows = []
            connect_point_cols = []

            # 扫描前10行查找"连接点"或"Connect Point"
            for row_idx in range(min(10, df.shape[0])):
                for col_idx in range(df.shape[1]):
                    cell_value = str(df.iloc[row_idx, col_idx]).strip().lower()
                    if "连接点" in cell_value or "connect point" in cell_value:
                        connect_point_rows.append(row_idx)
                        connect_point_cols.append(col_idx)
                        console_text.insert(tk.END,
                                            f"找到连接点标题: 第{row_idx + 1}行 第{col_idx + 1}列 - {cell_value}\n")

            # 然后在连接点位置下方查找"接点号"或"Terminal"
            for i, row_idx in enumerate(connect_point_rows):
                col_idx = connect_point_cols[i]

                # 检查连接点标题所在列及其右侧几列，查找接点号
                for c in range(col_idx, min(col_idx + 3, df.shape[1])):
                    # 检查连接点下方的2-3行
                    for r in range(row_idx + 1, min(row_idx + 4, df.shape[0])):
                        cell_value = str(df.iloc[r, c]).strip().lower()
                        if "接点号" in cell_value or "terminal" in cell_value:
                            terminal_columns.append(c)
                            console_text.insert(tk.END, f"找到接点号列: 第{r + 1}行 第{c + 1}列 - {cell_value}\n")

            # 如果没找到接点号列，尝试基于数据特征识别
            if not terminal_columns:
                console_text.insert(tk.END, "未找到明确的接点号列，尝试基于数据特征识别...\n")

                # 如果找到了连接点标题，检查它们下方的列
                if connect_point_cols:
                    for col_idx in connect_point_cols:
                        # 通常接点号会位于连接点标题右侧1-2列
                        potential_col = col_idx + 1
                        if potential_col < df.shape[1]:
                            # 检查这列是否包含大量数字(接点号通常是数字)
                            digit_count = 0
                            for r in range(10, min(30, df.shape[0])):  # 跳过可能的表头
                                cell_value = str(df.iloc[r, potential_col]).strip()
                                if cell_value.isdigit():
                                    digit_count += 1

                            # 如果超过半数是数字，可能是接点号列
                            if digit_count > 10:
                                terminal_columns.append(potential_col)
                                console_text.insert(tk.END,
                                                    f"基于数据特征识别到可能的接点号列: 第{potential_col + 1}列\n")

            # 如果仍然没找到，使用默认列
            if not terminal_columns:
                # 使用经验值：通常接点号在第3和第6列(索引2和5)
                default_cols = [2, 5]
                for c in default_cols:
                    if c < df.shape[1]:
                        terminal_columns.append(c)
                if terminal_columns:
                    console_text.insert(tk.END, f"使用默认接点号列: {[c + 1 for c in terminal_columns]}\n")
                else:
                    console_text.insert(tk.END, "未能识别接点号列，请检查Excel格式\n")
                    return

            # 去重
            terminal_columns = list(set(terminal_columns))
            console_text.insert(tk.END, f"最终使用的接点号列: {[col + 1 for col in terminal_columns]}\n")

            # 查找前缀并生成修改信息
            current_prefix = None
            modifications = []  # [(行索引, 列索引, 原值, 新值)]

            # 遍历每一行
            for row_idx in range(df.shape[0]):
                # 检查是否是前缀行
                for col_idx in range(df.shape[1]):
                    cell_value = str(df.iloc[row_idx, col_idx])
                    if re.match(r'^XS\d+$', cell_value.strip()):
                        current_prefix = cell_value.strip()
                        console_text.insert(tk.END, f"第{row_idx + 1}行: 检测到前缀 {current_prefix}\n")

                # 如果有前缀，处理接点号列
                if current_prefix:
                    for col_idx in terminal_columns:
                        if col_idx < df.shape[1]:  # 确保列索引有效
                            cell_value = str(df.iloc[row_idx, col_idx])
                            if cell_value != 'nan' and cell_value.strip() and (re.search(r'\d', cell_value) or re.search(r'[a-zA-Z]', cell_value)):#目前匹配需单元格内必须有数字，如果是纯字母就不行
                                # 检查是否需要添加前缀
                                if not re.search(fr'^{current_prefix}\.', cell_value):
                                    processed_value = process_terminal_with_prefix(cell_value, current_prefix)
                                    modifications.append((row_idx, col_idx, cell_value, processed_value))

            # 应用修改：基于文件类型选择最佳方法
            console_text.insert(tk.END, f"开始应用 {len(modifications)} 个修改...\n")

            # 尝试多种方法修改文件，以确保格式保留
            format_preserved = False

            # 方法1：尝试使用xlwings（优先，可以保留所有格式）
            if not format_preserved:
                try:
                    import xlwings as xw
                    console_text.insert(tk.END, "使用xlwings修改Excel文件...\n")

                    # 创建Excel应用实例
                    app = xw.App(visible=False)
                    try:
                        # 打开工作簿
                        wb = app.books.open(output_file)
                        sheet = wb.sheets['Sheet1']

                        # 应用修改
                        for mod in modifications:
                            row_idx, col_idx, old_val, new_val = mod
                            # xlwings索引从1开始
                            sheet.cells(row_idx + 1, col_idx + 1).value = new_val

                        # 保存并关闭工作簿
                        wb.save()
                        wb.close()
                        format_preserved = True
                        console_text.insert(tk.END, "使用xlwings成功修改文件并保存，所有格式已保留\n")
                    finally:
                        app.quit()

                except Exception as e:
                    console_text.insert(tk.END, f"使用xlwings修改失败: {str(e)}\n")

            # 方法2：尝试使用win32com（次优先，可以保留所有格式）
            if not format_preserved:
                try:
                    import win32com.client
                    console_text.insert(tk.END, "使用win32com修改Excel文件...\n")

                    # 创建Excel应用实例
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False

                    try:
                        # 打开工作簿
                        wb = excel.Workbooks.Open(os.path.abspath(output_file))
                        sheet = wb.Worksheets("Sheet1")

                        # 应用修改
                        for mod in modifications:
                            row_idx, col_idx, old_val, new_val = mod
                            # win32com索引从1开始
                            sheet.Cells(row_idx + 1, col_idx + 1).Value = new_val

                        # 保存并关闭工作簿
                        wb.Save()
                        wb.Close()
                        format_preserved = True
                        console_text.insert(tk.END, "使用win32com成功修改文件并保存，所有格式已保留\n")
                    finally:
                        excel.Quit()

                except Exception as e:
                    console_text.insert(tk.END, f"使用win32com修改失败: {str(e)}\n")

            # 方法3：根据文件类型使用openpyxl或pandas
            if not format_preserved:
                if file_ext == '.xlsx':
                    try:
                        # 使用openpyxl直接修改xlsx文件
                        console_text.insert(tk.END, "使用openpyxl修改Excel文件...\n")
                        wb = openpyxl.load_workbook(output_file)
                        ws = wb['Sheet1']

                        # 应用修改
                        for mod in modifications:
                            row_idx, col_idx, old_val, new_val = mod
                            # openpyxl索引从1开始
                            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                            cell.value = new_val

                        # 保存工作簿
                        wb.save(output_file)
                        format_preserved = True
                        console_text.insert(tk.END, "使用openpyxl成功修改文件并保存，大部分格式已保留\n")
                    except Exception as e:
                        console_text.insert(tk.END, f"使用openpyxl修改失败: {str(e)}\n")

                # 对于xls文件，使用pandas处理
                if not format_preserved and file_ext == '.xls':
                    try:
                        console_text.insert(tk.END, "使用专用方法处理XLS文件...\n")

                        # 尝试使用xlrd和xlwt库
                        try:
                            import xlrd
                            import xlwt
                            from xlutils.copy import copy

                            # 打开原始工作簿
                            rb = xlrd.open_workbook(output_file, formatting_info=True)
                            wb = copy(rb)
                            sheet = wb.get_sheet(0)  # 假设Sheet1是第一个工作表

                            # 应用修改
                            for mod in modifications:
                                row_idx, col_idx, old_val, new_val = mod
                                sheet.write(row_idx, col_idx, new_val)

                            # 保存工作簿
                            wb.save(output_file)
                            format_preserved = True
                            console_text.insert(tk.END, "使用xlutils成功修改文件并保存，大部分格式已保留\n")
                        except Exception as xlrd_error:
                            console_text.insert(tk.END, f"使用xlrd/xlwt修改失败: {str(xlrd_error)}\n")
                    except Exception as e:
                        console_text.insert(tk.END, f"处理XLS文件失败: {str(e)}\n")

            # 最终备选方案：使用pandas
            if not format_preserved:
                console_text.insert(tk.END, "使用pandas处理Excel文件...\n")

                # 应用修改到DataFrame
                for mod in modifications:
                    row_idx, col_idx, old_val, new_val = mod
                    df.iloc[row_idx, col_idx] = new_val

                # 确定输出文件格式
                if file_ext == '.xls':
                    # 如果原文件是xls，可能需要转换为xlsx来最大化格式保留
                    output_xlsx = os.path.splitext(output_file)[0] + '.xlsx'
                    df.to_excel(output_xlsx, sheet_name='Sheet1', index=False, header=False)
                    console_text.insert(tk.END, f"已转换为XLSX格式并保存为: {output_xlsx}\n")
                    console_text.insert(tk.END, "注意：由于XLS格式限制，转换为XLSX格式能更好保留数据\n")
                else:
                    # 如果是xlsx，直接保存
                    df.to_excel(output_file, sheet_name='Sheet1', index=False, header=False)
                    console_text.insert(tk.END, f"已使用pandas保存为: {output_file}\n")
                    console_text.insert(tk.END, "注意：使用pandas保存可能会丢失一些格式\n")

            console_text.insert(tk.END, "处理完成！\n")

        except Exception as e:
            console_text.insert(tk.END, f"处理Excel文件时出错: {str(e)}\n")
            console_text.insert(tk.END, traceback.format_exc())

    except Exception as e:
        console_text.insert(tk.END, f"处理Excel文件时出错: {str(e)}\n")
        console_text.insert(tk.END, traceback.format_exc())
        # endregion