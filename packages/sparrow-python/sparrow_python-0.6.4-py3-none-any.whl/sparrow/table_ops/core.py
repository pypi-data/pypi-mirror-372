from __future__ import annotations
import pandas as pd
from typing import List, Union
import os
import hashlib

def groupby_choice(df: pd.DataFrame, by: Union[str, List], col_name: any, choice='max', inplace=True):
    """
    取分组后的某列最值,组成的新df. 默认inplace.

    Example::
     df = pd.DataFrame({'key' : ['A', 'A', 'B', 'B', 'C', 'C'],
                       'value' : ['v1', 'v2', 'v3', 'v4','v5', 'v6'],
                       'prob' : [1, 5, 50, 2, 5, 5]})
    >>> df
        key value  prob
    0   A    v1     1
    1   A    v2     5
    2   B    v3    50
    3   B    v4     2
    4   C    v5     5
    5   C    v6     5
    >>> groupby_choice(df, 'key', 'prob', 'max')
    >>>
        key value  prob
    1   A    v2     5
    2   B    v3    50
    4   C    v5     5
    """
    if not inplace:
        df = df.copy(deep=True)
    index_list = []
    for idx, item in df.groupby(by)[col_name]:
        if choice == "max":
            index_list.append(item.idxmax())
        elif choice == "min":
            index_list.append(item.idxmin())
        else:
            raise "Invalid `func` parameter."
    return df.iloc[index_list]


def group_df(df, col_name, interval=5, use_max_min_interval=False, closed='neither', dropna=True):
    """
    Parameters
    ----------
        col_name: 根据 `col_name` 进行分组
        interval: 合并采样间隔
        use_max_min_interval: True使用最大最小区间确定等距采样个数； False使用df的样本数目确定采样个数

    """
    if dropna:
        df = df.dropna(axis=0, how='any', inplace=False)
    df = df.sort_values(by=col_name, ascending=True)
    if use_max_min_interval:
        periods = (df[col_name].max() - df[col_name].min()) / interval
    else:
        periods = len(df) // interval

    bins = pd.interval_range(df[col_name].min(), df[col_name].max(),
                             periods=periods,
                             closed=closed)
    pd_cut = pd.cut(df[col_name], bins=bins)
    for idx, i in enumerate(df.groupby(pd_cut)):
        agg_res = i[1].agg('mean')
        if idx == 0:
            df_grouped = agg_res
        else:
            df_grouped = pd.concat([df_grouped, agg_res], axis=1)
    df_grouped = df_grouped.transpose()
    return df_grouped.dropna().reset_index(inplace=False).drop(['index'], axis=1)


def re_ord_df_col(df, col_name, ord_num=0):
    """Re-order df's column name."""
    tmp_list = df.columns.tolist()
    tmp_list.remove(col_name)
    tmp_list.insert(ord_num, col_name)
    df = df[tmp_list]
    return df


def guess_str_fmt(time_str: str, token: str):
    time_list = time_str.split(token)
    list_len = len(time_list)
    if list_len == 3:
        return f"%Y{token}%m{token}%d"
    elif list_len == 2:
        return f"%Y{token}%m"
    elif list_len == 1:
        if len(time_str) == 4:
            return f"%Y"
        elif len(time_str) == 6:
            return f"%Y%m"
        elif len(time_str) == 8:
            return f"%Y%m%d"
        else:
            return None
    else:
        raise ValueError("Invalid datetime format.")


def guess_datetime_fmt(timeseries: List[str], token_list=('-', '/', ' ', '_', '.')):
    """Guess datetime format."""
    for token in token_list:
        time_format = guess_str_fmt(timeseries[0], token)
        if time_format:
            break
    else:
        raise ValueError("Invalid datetime format.")
    return time_format


def insert_line(df: pd.DataFrame, idx, new_line: pd.Series | pd.DataFrame | dict, ignore_index=True):
    df_head = df.iloc[:idx, :]
    df_tail = df.iloc[idx:, :]
    if isinstance(new_line, dict):
        df_line = pd.DataFrame(new_line)
    elif isinstance(new_line, pd.Series):
        df_line = pd.DataFrame(new_line).T
    else:
        df_line = new_line
    df_new = pd.concat([df_head, df_line, df_tail], ignore_index=ignore_index).reset_index(drop=True)
    return df_new


def extract_excel_with_images(
    excel_path,
    image_column_name,
    image_output_dir="extracted_images",
    sheet_name=0,
    use_hash_filename=False,
    save_updated_excel=False,
    image_output_dir_prefix=False,
):
    """
    读取包含图片的Excel文件，提取图片并将其与表格行对齐。

    参数：
    - excel_path: Excel文件路径
    - image_column_name: 图片所在的列名称
    - image_output_dir: 图片保存路径
    - sheet_name: 指定读取的工作表（默认读取第一个）
    - use_hash_filename: 是否使用哈希值作为文件名（默认为False）
    - save_updated_excel: 是否保存更新后的Excel文件（默认为False）
    - image_output_dir_prefix: 图片输出目录前缀, 如果为False，则不使用前缀, 只有文件名，不包含路径

    返回：
    - 一个带有图片路径的新DataFrame
    """

    from openpyxl import load_workbook
    # 获取Excel文件名（不含路径和扩展名）
    excel_basename = os.path.splitext(os.path.basename(excel_path))[0]

    # 加载 Excel 文件
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.active

    # 如果sheet_name为None或数字，获取当前活动的工作表名
    actual_sheet_name = sheet_name if isinstance(sheet_name, str) else ws.title

    # 创建特定于此Excel的图片输出目录
    # 如果只有一个工作表，不创建额外的工作表子目录
    if len(wb.worksheets) == 1:
        specific_image_dir = os.path.join(image_output_dir, excel_basename)
    else:
        specific_image_dir = os.path.join(
            image_output_dir, excel_basename, actual_sheet_name
        )
    os.makedirs(specific_image_dir, exist_ok=True)

    # 使用openpyxl直接读取数据
    data = []
    headers = []
    
    # 获取表头
    for cell in ws[1]:
        headers.append(cell.value)
    
    # 获取所有数据行
    for row in ws.iter_rows(min_row=2):  # 从第二行开始（跳过表头）
        row_data = []
        for cell in row:
            row_data.append(str(cell.value) if cell.value is not None else "")
        data.append(row_data)
    
    # 创建DataFrame
    df = pd.DataFrame(data, columns=headers)
        
    # 打印DataFrame信息
    print(f"DataFrame信息:")
    print(f"- 总行数: {len(df)}")
    print(f"- 列名: {list(df.columns)}")
    print(f"- 数据类型:\n{df.dtypes}")

    
    # 确保指定的列名存在
    if image_column_name not in df.columns:
        raise ValueError(f"列名 '{image_column_name}' 在Excel文件中不存在")

    # 获取列名对应的列索引（从1开始）
    cols = list(df.columns)
    image_column = cols.index(image_column_name) + 1

    # 提取图片和位置
    image_map = {}  # row_num: image_filename
    print(f"开始提取图片，总行数: {len(df)}")

    
    for img in ws._images:
        anchor = img.anchor._from  # 起始锚点
        row = anchor.row + 1
        col = anchor.col + 1

        if col != image_column:
            continue  # 忽略不在目标列的图片

        if use_hash_filename:
            # 使用图片内容生成哈希值作为文件名
            img_hash = hashlib.md5(img.ref.getvalue()).hexdigest()
            filename = f"{img_hash}.png"
        else:
            filename = f"image_r{row}_c{col}.png"

        save_path = os.path.join(specific_image_dir, filename)
        # 保存图片
        with open(save_path, "wb") as img_file:
            img_file.write(img.ref.getvalue())

        # 根据 image_output_dir_prefix 参数决定保存的路径格式
        if image_output_dir_prefix:
            # 使用相对路径
            image_map[row] = os.path.join(specific_image_dir, filename)
        else:
            # 只使用文件名
            image_map[row] = filename

    print(f"找到的图片数量: {len(image_map)}")

    # 添加图片路径到原始图像列
    for row_num, img_path in image_map.items():
        df_index = row_num - 2  # 减2：1行表头 + 1行从1开始变成从0
        if 0 <= df_index < len(df):
            df.loc[df_index, image_column_name] = img_path

    # 生成更新后的Excel文件
    if save_updated_excel:
        output_excel_path = os.path.join(
            os.path.dirname(excel_path),
            f"{excel_basename}_updated{os.path.splitext(excel_path)[1]}",
        )
        df.to_excel(output_excel_path, sheet_name=actual_sheet_name, index=False)
        print(f"更新后的Excel文件已保存到: {output_excel_path}")

    return df


if __name__ == "__main__":
    ts = ['2022-01', '2022/02']
    df = pd.DataFrame({'date': ts, })
    time_format = guess_datetime_fmt(ts)
    print(time_format)
    df['date'] = pd.to_datetime(ts)
    print(df)
