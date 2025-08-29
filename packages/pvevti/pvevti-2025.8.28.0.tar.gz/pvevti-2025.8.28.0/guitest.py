"""
TODO: 
 - Multithread! @Dave M. good luck!
"""

from PySide6 import QtWidgets, QtGui
# from PySide6.QtCore import QRunnable, Qt, QThreadPool
from PySide6.QtWidgets import (QWidget, QSplashScreen, QFileDialog, 
                               QListWidgetItem, QProgressBar, 
                               QApplication, QDialog, QDialogButtonBox)
from PySide6.QtGui import QPixmap, QKeySequence, QShortcut, QGuiApplication
from PySide6.QtCore import Qt, QObject, Signal, QThread
import sys, os
from getpass import getuser
from certifi import where
from tempfile import gettempdir
from time import sleep
from importlib.metadata import version
from matplotlib.backends.backend_qtagg import FigureCanvas
from matplotlib.figure import Figure
from matplotlib import pyplot as plt
import geopandas as gpd
import contextily as ctx
import requests
from shapely.geometry import Point
from pvevti import pdfutil

# Yield versions of dependencies
pvevti_version =    version('pvevti')
PySide6_version =   version('PySide6')
FPT_version    =    '2025.8.20'

# Request network access and set caches
cache_dir = os.path.join(os.getcwd(), 'cache')
os.environ['REQUESTS_CA_BUNDLE'] =  where()
os.environ['TMPDIR'] =              gettempdir()
os.environ['XDG_CACHE_HOME'] =      cache_dir
os.environ["QT_API"] =              "PySide6"

# Create directory for caching if it doesn't exist yet
os.makedirs(cache_dir, exist_ok=True)

# Detect a frozen runtime (as when run as an executable) and readdress two resource nodes
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
    os.environ['GDAL_DATA'] = os.path.join(base_path, 'rasterio', 'gdal_data')
    os.environ['PROJ_LIB'] = os.path.join(base_path, 'pyproj', 'share', 'proj')

# Preset a few paths
username =      getuser() # Username of the current active user
default_path =  "C:/Users/"+username+"/Downloads/" # Default path if none are known
dirs_path =     "C:/Users/"+username+"/AppData/Local/DT-FPT/" # Location in %Appdata% which stores the directories
basedir =       os.path.dirname(__file__) # Base directory for use in finding resources (icons)
pdfutil.IMAGE_CACHE_DIR = dirs_path # Enable offline image caching for report creation

# Preload default path preferences, to be changed later.
csv_open_default_path = default_path
csv_save_default_path = default_path
rep_save_default_path = default_path
runlog =                default_path
runlog_dir =            default_path
combine_prefs =         default_path

# Instantiate lists for custom report paths and names
custom_reports =        []
custom_report_names =   []

# Ensure directories exist
os.makedirs(dirs_path, exist_ok=True)

# Provide a path irrespective of unbuilt/built status (Courtesy of OpenAI's ChatGPT)
def resource_path(relative_path):
    if getattr(sys, 'frozen', False):  # Running in a PyInstaller bundle
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Create directories txt if it doesn't exist
try:
    with open(dirs_path+'/dirs.txt', "x") as f:
        pass
except FileExistsError:
    print(f"File '{dirs_path+'/dirs.txt'}' already exists!")

# Get last opened directories
with open(dirs_path+'/dirs.txt') as file:
    lines = file.read().split('\n')
    for line in lines:
        path = ':'.join(line.split(":")[1:]).strip()
        print(line, '->',path)
        if 'LASTCSVGETPATH' in line:
            csv_open_default_path = path
        elif 'LASTCSVSAVEPATH' in line:
            csv_save_default_path = path
        elif 'LASTREPORTSAVEPATH' in line:
            rep_save_default_path = path
        elif 'RUNLOG' in line:
            runlog = path
        elif 'RUNLOGDIR' in line:
            runlog = path
        elif 'COMBINEPREFS' in line:
            combine_prefs = path
        elif 'CUSTOMREPORT_' in line:
            name = line.split(":")[0].replace("CUSTOMREPORT_", "")
            if path not in custom_reports:
                print(name, ":", path)
                custom_report_names.append(name)
                custom_reports.append(path)

# Default reports and user-supplied reports
reports = ["TestConditions", "PowerSteering", "Cooling", "Overview", "Electrical", "HVAC", "TowDyno"]
reports += custom_report_names


# sets the specified default str to the chosen path.
def configure_default_path(name:str, path:str):
    print(" OPEN: ",str(path))
    print(" FILE: ",str(dirs_path+'/dirs.txt'))
    with open(dirs_path+'/dirs.txt', 'r') as file:
        lines = file.readlines()
    i = -1
    for (idx, line) in enumerate(lines):
        if name in line:
            i = idx
    
    if i != -1:
        print("Found",i)
        lines[i] = name+":"+path
    else:
        print("Unfound")
        print(lines,end='')
        lines.append(name+":"+path)
        print(" ->",lines)
    with open(dirs_path+'/dirs.txt', 'w') as file:
        file.writelines('\n'.join([item for item in lines if item != '\n']))

def clear_custom_paths():
    print(" FILE: ",str(dirs_path+'/dirs.txt'))
    with open(dirs_path+'/dirs.txt', 'r') as file:
        lines = file.readlines()
    len1 = len(lines)
    lines = [line for line in lines if "CUSTOMREPORT_" not in line]
    print(" > Deleted {} configs".format(len1 - len(lines)))
    with open(dirs_path+'/dirs.txt', 'w') as file:
        file.writelines('\n'.join([item for item in lines if item != '\n']))

# Color widget, filled with the specified solid color
class Color(QWidget):
    def __init__(self, color):
        from PySide6.QtGui import QColor, QPalette
        super().__init__()
        self.setAutoFillBackground(True)

        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Window, QColor(color))
        self.setPalette(palette)

# Fix for python override on taskbar icon
try:
    from ctypes import windll
    myappid = 'com.dtna.fpt'
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except ImportError:
    pass


# Styles
titleStyle =        'font-size: 20px; font-family: Aptos; padding: 8px 0px 0px 8px; font-weight:bold; color:black; background-color:#E6E6E6;'
subtitleStyle =     'font-size: 12px; font-family: Aptos; padding: 0px 0px 0px 12px; color:black; background-color:#E6E6E6;'
toolbarButtonStyle =    """
    QToolBar {
        border: 0px;
        background-color: #00677F;
        spacing : 4px;
    }
    QToolBar QToolButton{
        color: white;
        font-size : 11px;
        border-radius: 0px;
        margin: 0px 5px 0px 0px;
        font-family: Aptos;
    }
    QToolBar QToolButton:hover{
        background-color: #307f91;
    }
"""
prefsTitleStyle = """
    QLabel {
        margin: 8px 0px 0px 0px;
        font-size: 12px;
        font-weight: bold;
        color:black;
    }
"""
smallText = """font-size: 10px; padding: 2px 2px 2px 6px; color:black;background-color:#F6F6F6;"""
smallTextIndent = """font-size: 12px; margin: 0px 0px 0px 10px; color:black;"""
smallIndent = """margin: 12px 0px 0px 6px; color:black;"""
checkBoxStyle = """
QCheckBox {
        spacing: 10px;
        font-size: 14px;
        font-family: 'Segoe UI';
        color: #333;
    }

    QCheckBox::indicator {
        width: 10px;
        height: 10px;
        border-radius: 3px;
        border: 1px solid #DDD;
        background-color: #F6F6F6;
    }

    QCheckBox::indicator:checked {
        background-color: #5097AB;
        border: 1px solid #A6CAD8;
    }

    QCheckBox::indicator:unchecked {
        background-color: #f0f0f0;
        border: 1px solid #ccc;
    }

    QCheckBox::indicator:hover {
        border: 1px solid #00677F;
        background-color: #00677F;
    }

    QCheckBox::indicator:checked:hover {
        background-color: #004355;
    }
"""
buttonStyle = """
QPushButton{
    margin: 0px 0px 0px 0px;
    padding: 2px 8px 2px 8px;
    color:#333;
    font-size:10px;
    background-color:#F6F6F6;
}
QPushButton:pressed {
    border: 1px inset #005FBA;
    padding: 1px 7px 1px 7px;
    margin: 0px 0px 0px 0px;
    border-radius: 5px;
    color:#F6F6F6;
    background-color:#5097AB;
}
"""
listwidgetStyleTight = """
    QListWidget {
        margin: 0px;
        padding:0px;
        background-color:#F6F6F6;
        border: none;
        font-size:11px;
    }
    QListWidget::item {
        margin: 0px;
        padding: 0px;
        font-size:11px;
        color: black;
    }
    QListWidget::item:hover {
        background-color: #5097AB;
        color: black;
    }
    QListWidget::item:selected {
        background-color: #00677F;
        color: #F0F0F0;
    }
    QScrollBar:vertical {
        border: none;
        background: #E8E8E8;
        width: 6px;
        margin: 0px 0px 0px 0px;
    }
    QScrollBar::handle:vertical {
        background: #D0D0D0;
        min-height: 15px;
        border-radius: 0px;
    }
    QScrollBar::handle:vertical:hover {
        background: #B0B0B0;
    }
    QScrollBar::handle:vertical:pressed {
        background: #929292;
    }
    QScrollBar::add-line:vertical {
        border: none;
        background: none;
    }
    QScrollBar::sub-line:vertical {
        border: none;
        background: none;
    }
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
        background: none;
    }
"""
listwidgetStyle = """
    QListWidget {
        margin: 2px;
        background-color:#F6F6F6;
        border: none;
    }
    QListWidget::item {
        margin: 0px;
        padding: 3px;
        color: black;
    }
    QListWidget::item:hover {
        background-color: #5097AB;
        color: black;
    }
    QListWidget::item:selected {
        background-color: #00677F;
        color: #F0F0F0;
    }
    QScrollBar:vertical {
        border: none;
        background: #E8E8E8;
        width: 8px;
        margin: 0px 0px 0px 0px;
    }
    QScrollBar::handle:vertical {
        background: #D0D0D0;
        min-height: 20px;
        border-radius: 0px;
    }
    QScrollBar::handle:vertical:hover {
        background: #B0B0B0;
    }
    QScrollBar::handle:vertical:pressed {
        background: #929292;
    }
    QScrollBar::add-line:vertical {
        border: none;
        background: none;
    }
    QScrollBar::sub-line:vertical {
        border: none;
        background: none;
    }
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
        background: none;
    }
"""
iconButtonStyle = """
    QPushButton {
        margin: 0px;
        border-radius:0px;
        border:none;
        background-color:#E8E8E8;
        padding: 0px;
        width: 36px;
        height:36px;
    }
    QPushButton:hover {
        background-color: #5097AB;
    }
    QPushButton:pressed {
        background-color: #00677F;
    }
"""
statusLabelStyle = """
font-size:10px;
margin:1px 1px 8px 10px;
font-family: "Courier New";
"""
viewBtnStyle = """
QPushButton{
    font-size:10px;
    text-decoration: underline;
    border: none;
    color: black;
    margin: 2px 2px 2px 2px;
}
QPushButton:hover {
    color: #5097AB;
}
QPushButton:pressed {
    color: #00677F;
}
"""
progressBarStyle = """
QProgressBar {
    border: none;
    border-radius: 5px;
    text-align: center;
    font-size:10px;
    padding:0px 0px 0px 0px;
    background-color: #E8E8E8;
    color: black;
    margin:0px 0px 0px 2px;
}
QProgressBar::chunk {
    background-color: #5097AB;
}
"""
progressBarStyleThin = """
    QProgressBar {
        border: none;
        border-radius: 5px;
        text-align: center;
        font-size:10px;
        padding:0px 0px 0px 0px;
        background-color: #E8E8E8;
        color: black;
        margin:0px 0px 0px 0px;
    }
    QProgressBar::chunk {
        background-color: #5097AB;
    }
"""
mf4DialogTitle = 'font-size:20px;font-weight:bold;color:black;margin:0px 0px 0px 10px;'
mf4DialogProceed = "QPushButton:Enabled{background-color:#D0D0D0;color:black;border:none;border-radius:0px;margin:0px 0px 0px 10px;} QPushButton:Enabled::hover{background-color:#F0F0F0;} QPushButton:Disabled{background-color:#E6E6E6;color:#D0D0D0;border:1px solid #D0D0D0;border-radius:0px;}"
mainPane = "background-color: #E6E6E6;"
buttonParentStyle = 'background-color:#E6E6E6;margin:0px 0px 0px 12px;'

# Globals for keeping track of things during runtime
savedir_csv = ""
savedir_rep = ""

# Reserve for paths and check boxes
paths = []
boxes = []
filterSignals = []
SLC_Date = ""
MAX_CORES = 6

# Reserve for preferences
doExport = True
replaceOriginals = False

# Check for internet on load
try: 
    requests.get('http://www.google.com', timeout=5)
    internet = True
except requests.ConnectionError:
    internet = False
except requests.Timeout:
    internet = False

class workerSignals (QObject):
    finished = Signal()

class Worker(QObject):
    def __init__(self):
        super().__init__()
        self.signals = workerSignals
        

# Load images into appdata
def load_images(splash, app):
    # TODO: Load stuff into appdata
    from pvevti.genutil import geoFence, getGPSBox
    from time import sleep
    imgs = [x for x in os.listdir(dirs_path) if '.tiff' in x]
    imgs_names = [x.replace('.tiff','') for x in imgs]
    print("Images:",imgs)
    to_download = []
    # Loop through all tracked routes from pvevti.genutil
    for route in geoFence:
        # Update splash
        splash.showMessage("Pre-cache: {}".format(route['name']), Qt.AlignBottom | Qt.AlignHCenter, Qt.white)
        app.processEvents()
        # Compile routes needed to download
        if route['name'] not in imgs_names:
            to_download.append(route)

    # Threading NOT COMPLETE
    # Currently, the code under if internet: runs on a single thread, which is also the GUI's thread.
    # TODO:
    #   For each core to use, assign routes to download. The to_download list contains ALL routes to download.
    #   For each core, start a downloader worker process and assign it to a core object. Look into PySide6 multithreading to see more info.
    #   Have each core work through their respective tasks, and emit a complete boolean signal when done.
    #   Once all cores are done, start UI.
    cores_to_use = min(max(0, len(to_download)), MAX_CORES)
    threads = []
    workers = []
    for i in range(0, cores_to_use):
        print("  [Download] Starting core {}".format(i+1))
        print("    Responsible to download: {}".format(to_download[i:cores_to_use:-1]))


    # Download routes:
    if internet:
        for idx, route in enumerate(to_download):
            # Try to download
            try:
                [west, south, east, north] = getGPSBox(route['box'])
                mid_x = west + (east - west) / 2
                mid_y = south + (north - south) / 2
                span = max(north-south, east-west)
                span = max(min(span * 1.1, span + 0.1), 0.01)
                west = mid_x - span / 2
                east = mid_x + span / 2
                south = mid_y - span / 2
                north = mid_y + span / 2
                name = route['name']
                splash.showMessage("Download: {} ({}/{})".format(route['name'], idx+1, len(to_download)), Qt.AlignBottom | Qt.AlignHCenter, Qt.white)
                app.processEvents()
                ctx.bounds2raster(west, south, east, north,
                                    ll=True, path=dirs_path+name+".tiff", 
                                    source=ctx.providers.OpenStreetMap.Mapnik)
                print("Added "+name+".tiff to "+dirs_path)
            except:
                print("Failed to pre-cache {}.tiff".format(name))
    else:
        print("No internet!\n {} Files skipped download.".format(len(to_download)))
        splash.showMessage("No Internet", Qt.AlignBottom | Qt.AlignHCenter, Qt.white)
        app.processEvents()
        sleep(1)

# Create a plot popup
class Plot(FigureCanvas):
    def __init__(self, parent=None, width=400, height=400):
        dpi = 50
        width = width / dpi
        height = height / dpi
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.subplots_adjust(left=0,right=1,top=1,bottom=0)
        self.axes = fig.add_subplot(111)
        super().__init__(fig)

# TESTING: Map Popup
class MapPopup(QDialog):
    def __init__(self, parent = None, filename=""):
        super().__init__(parent)
        self.setWindowFlags(Qt.Window | Qt.FramelessWindowHint)

        from PySide6.QtWidgets import QVBoxLayout, QLabel

        closeSC = QShortcut(QKeySequence('Space'), self)
        closeSC.activated.connect(self.close)


        layout = QVBoxLayout()
        layout.setContentsMargins(0,0,0,0)

        if os.path.isfile(filename) and '.csv' in filename:
            print("Map Generation: "+filename)
            from pvevti.csvutil import df_from_csv
            from pvevti.genutil import gps_filter_data, getRoute, geoFence, getGPSBox
            from pandas import DataFrame
            
            try:
                df = df_from_csv(filename, ['GPS_x[°]', 'GPS_y[°]']) # Create dataframe
                df_length = df.shape[0] # Get length of DF
                target_length = 500 # Target length for processing acceleration

                # Sampling interval - take every (interval)th sample to attain ~500 total datapoints
                interval = max(1, int(df_length/target_length))
                x_data = df['GPS_x[°]'].tolist()[::interval]
                y_data = df['GPS_y[°]'].tolist()[::interval]
                df = DataFrame({'GPS_x[°]':x_data, 'GPS_y[°]':y_data})

                # Extract route from microDF
                route=getRoute(df)
                print(geoFence,'\n[{}] {}'.format(route[0], route[1]))

                # Find matching route
                routeNames = [rt['name'] for rt in geoFence]
                if route[0] in routeNames:
                    # Found state
                    currentRoute = geoFence[routeNames.index(route[0])]
                    print("Current Route: "+str(currentRoute))
                    [west, south, east, north] = getGPSBox(currentRoute['box']) # get bounding box of route, squarify, and calculate span + mid variables
                    span = max(east-west, north-south)
                    span = max(min(span * 1.1, span + 0.1), 0.01)
                    mid_x = west + (east - west) / 2
                    mid_y = south + (north - south) / 2
                else: 
                    # Unfound state
                    currentRoute = {'name': 'UK', 'desc': 'Unknown'}

                    # Get bounding box of data, calculate span + mid variables
                    max_x = max(x_data)
                    min_x = min(x_data)
                    max_y = max(y_data)
                    min_y = min(y_data)
                    span = max(max_x-min_x, max_y-min_y)
                    span = max(min(span * 1.1, span + 0.1), 0.01)
                    mid_x = (max_x-min_x)/2 + min_x
                    mid_y = (max_y-min_y)/2 + min_y

                print("DF Length: {}\nx_data Length: {}".format(df_length, len(x_data)))

                sc = Plot(self)
                
                sc.axes.plot(x_data, y_data)
                sc.axes.text(0.5, 0.95, '{} ({})'.format(currentRoute['name'], currentRoute['desc']), 
                             horizontalalignment='center', verticalalignment='center', 
                             transform=sc.axes.transAxes, fontsize=20, bbox=dict(facecolor='white', alpha=0.3, linewidth=0))
                
                sc.axes.set_axis_off() # Remove axis
                
                sc.axes.set_xlim([mid_x-span/2, mid_x+span/2])
                sc.axes.set_ylim([mid_y-span/2, mid_y+span/2])
                
                path = dirs_path + currentRoute['name'] + '.tiff' # Set path (theoretical location)
                if os.path.isfile(path):
                    print("Pull local file: "+path)
                    ctx.add_basemap(sc.axes, crs='EPSG:4326', source=path) # Load from disk
                else:
                    print("Download file: "+currentRoute['name'])
                    ctx.add_basemap(sc.axes, crs='EPSG:4326', source=ctx.providers.OpenStreetMap.Mapnik) # Download

                layout.addWidget(sc) # Add the plot to the window

            except Exception as e:
                print("Failure to extract data!\n {}".format(str(e)))
        
        self.setLayout(layout)

# Informational Popup
class InfoPopup(QDialog):
    def __init__(self, parent=None, windowTitle:str="Unnamed Window", windowSubtitle:str="List of example signals", windowContents:list=['Signal A', 'Signal B', 'Signal C']):
        super().__init__(parent=parent)

        windowContents = "\n".join(windowContents)

        from PySide6.QtWidgets import QLabel, QPushButton, QVBoxLayout
        self.setWindowTitle(windowTitle)
        layout = QVBoxLayout()
        label = QLabel(windowSubtitle)
        label.setStyleSheet("""
                            qproperty-alignment: AlignHCenter;
                            font-size: 18px;
                            """)
        content = QLabel(windowContents)
        content.setStyleSheet("font-size:11px;")
        openReport = QPushButton('Open Report')
        openReport.clicked.connect(self.accept)

        layout.addWidget(label)
        layout.addWidget(content)
        layout.addWidget(openReport)
        self.setLayout(layout)

# MF4 Processing Dialog
class MF4Dialog(QDialog):
    def __init__(self, parent=None):
        self.mf4_path = ""
        self.MF4obj = None
        self.searchResults = []
        self.selectedResults=[]
        self.search_signals =None
        search_style = """
            QTextEdit,QPlainTextEdit{
                color:black;
                background-color:#F6F6F6;
                border:none;
                margin: 0px 0px 0px 0px;
                padding: 1px 1px 1px 1px;
            }
            QTextEdit:focus,QPlainTextEdit:focus{
                border-left: 2px solid #00677F;
                background-color:#FAFAFC;
                padding: 1px 1px 1px 5px;
            }"""
        super().__init__(parent=parent)
        self.setStyleSheet("""
                           color:black; 
                           background-color:#E6E6E6;
                           """)
        from PySide6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QPlainTextEdit, QTextEdit, QListWidget, QProgressBar
        wrapperWidget = QWidget(parent=parent)
        wrapperWidget.setFixedSize(800, 360)
        self.setWindowTitle('MF4 Dialog')
        toplevel_layout = QVBoxLayout()
        toplevel_layout.setContentsMargins(0, 5, 0, 0)
        title = QLabel('MF4 Processing Dialog')
        title.setStyleSheet(mf4DialogTitle)
        toplevel_layout.addWidget(title)

        file_select_layout = QHBoxLayout()
        pick_btn = QPushButton('Pick File')
        pick_btn.setStyleSheet(str(mf4DialogProceed))
        pick_btn.setFixedSize(60, 24)
        pick_btn.clicked.connect(self.pick_file)
        self.text_entry = QPlainTextEdit()
        self.text_entry.setPlaceholderText('Use the "Pick File" button or enter a path here')
        self.text_entry.setFixedHeight(26)
        self.text_entry.setStyleSheet(search_style)
        self.text_entry.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.text_entry.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.text_entry.textChanged.connect(self.check_open)
        file_select_layout.addWidget(pick_btn)
        file_select_layout.addWidget(self.text_entry)
        file_select_layout.addSpacing(15)
        toplevel_layout.addLayout(file_select_layout)
        toplevel_layout.addSpacing(4)
        bottom_container = QWidget()
        # bottom_container.setFixedHeight(300)
        bottom_container.setStyleSheet('margin:0px 5px 0px 0px;')

        bottom_layout = QHBoxLayout()
        search_container = QWidget()
        # search_container.setStyleSheet("background-color:#E6E6E6;")
        search_container.setFixedWidth(320)
        search_layout = QVBoxLayout()
        search_layout.setContentsMargins(5, 3, 5, 3)

        self.nameSearch = QTextEdit()
        self.nameSearch.setFixedHeight(24)
        self.nameSearch.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.nameSearch.setPlaceholderText('Search by channel name')
        self.nameSearch.textChanged.connect(self.updateListFromName)
        self.nameSearch.setStyleSheet(search_style)
        self.nameSearch.setToolTip('Enter a channel name to search.\nUse "=" to find exactly the search string.\nNot case sensitive.')

        self.unitSearch = QTextEdit()
        self.unitSearch.setFixedHeight(24)
        self.unitSearch.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.unitSearch.setPlaceholderText('Search by channel unit')
        self.unitSearch.textChanged.connect(self.updateListFromUnit)
        self.unitSearch.setStyleSheet(search_style)
        self.unitSearch.setToolTip('Enter a unit to search.\nUse "/deg" to search for degrees.\nNot case sensitive.')

        self.sourceSearch = QTextEdit()
        self.sourceSearch.setFixedHeight(24)
        self.sourceSearch.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sourceSearch.setPlaceholderText('Search by channel group source')
        self.sourceSearch.textChanged.connect(self.updateListFromSource)
        self.sourceSearch.setStyleSheet(search_style)
        self.sourceSearch.setToolTip('Enter a source to search. Not case sensitive.')

        self.signalList = QListWidget()
        self.signalList.itemDoubleClicked.connect(self.changeList)
        self.signalList.setStyleSheet(listwidgetStyleTight)
        search_layout.addWidget(self.nameSearch)
        search_layout.addWidget(self.unitSearch)
        search_layout.addWidget(self.sourceSearch)
        search_layout.addWidget(self.signalList)
        # search_layout.addStretch(1)
        search_container.setLayout(search_layout)
        list_container = QWidget()
        # list_container.setStyleSheet("background-color:#E6E6E6;")
        list_layout = QVBoxLayout()
        list_layout_header = QHBoxLayout()
        list_layout_header.addWidget(QLabel('Selected Signals'))
        list_layout_header.addStretch(1)
        self.export_search = QPushButton('Export selected to file')
        self.export_search.clicked.connect(self.exportSearch)
        self.export_search.setStyleSheet(viewBtnStyle)
        self.export_search.setVisible(False)
        self.add_all = QPushButton('Add all from search')
        self.add_all.clicked.connect(self.addAll)
        self.add_all.setStyleSheet(viewBtnStyle)
        self.add_all.setVisible(False)
        add_from_file = QPushButton('Import from file')
        add_from_file.clicked.connect(self.addFromFile)
        add_from_file.setStyleSheet(viewBtnStyle)
        clear_all = QPushButton('Clear')
        clear_all.clicked.connect(self.clearAll)
        clear_all.setStyleSheet(viewBtnStyle)
        list_layout_header.addWidget(self.export_search)
        list_layout_header.addWidget(add_from_file)
        list_layout_header.addWidget(self.add_all)
        list_layout_header.addWidget(clear_all)
        list_layout.addLayout(list_layout_header)
        self.selectedList = QListWidget()
        self.selectedList.itemDoubleClicked.connect(self.changeSelected)
        self.selectedList.setStyleSheet(listwidgetStyleTight)
        self.selectedList.setFixedHeight(180)
        list_layout.addWidget(self.selectedList)
        bottom_section = QHBoxLayout()
        self.proceed = QPushButton('Proceed')
        self.proceed.setFixedSize(80, 24)
        self.proceed.setEnabled(False)
        self.proceed.setStyleSheet(mf4DialogProceed)
        self.proceed.clicked.connect(self.execute)
        self.sampleFreq = QTextEdit()
        self.sampleFreq.setPlaceholderText('Sampling Frequency (default 1s)')
        self.sampleFreq.setFixedHeight(24)
        self.sampleFreq.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sampleFreq.setStyleSheet(search_style)
        self.sampleFreq.setToolTip('Use any of the following units or provide none to default to seconds: h, m, s, ms, us. Ex: 250ms or 25m')
        bottom_section.addWidget(self.sampleFreq)
        bottom_section.addStretch(1)
        bottom_section.addWidget(self.proceed)
        list_layout.addLayout(bottom_section)
        list_layout.setContentsMargins(0,0,16,0)
        list_container.setLayout(list_layout)
        bottom_layout.addSpacing(6)
        bottom_layout.addWidget(search_container)
        bottom_layout.addWidget(list_container)
        bottom_layout.setContentsMargins(0,0,0,0)
        bottom_container.setLayout(bottom_layout)
        toplevel_layout.addWidget(bottom_container)
        self.status = QLabel('Status: Idle')
        self.status.setStyleSheet('margin:0px 0px 0px 15px;')
        toplevel_layout.addWidget(self.status)
        self.statusBar = QProgressBar()
        self.statusBar.setMaximum(1)
        self.statusBar.setMinimum(0)
        self.statusBar.setValue(0)
        self.statusBar.setTextVisible(False)
        self.statusBar.setStyleSheet(progressBarStyleThin)
        self.statusBar.setFixedHeight(5)
        toplevel_layout.addWidget(self.statusBar)
        wrapperWidget.setLayout(toplevel_layout)
        wrapperLayout = QVBoxLayout()
        wrapperLayout.setContentsMargins(0, 0, 0, 0)
        self.setFixedSize(800, 360)
        wrapperLayout.addWidget(wrapperWidget)
        self.setLayout(wrapperLayout)
    
    def updateState(self, state:str):
        """
        Updates the status to the provided string and processes events.
        """
        self.status.setText('Status: '+state)
        app.processEvents()

    def execute(self):
        """
        
        """
        try:
            self.statusBar.setMaximum(0)
            frq = self.sampleFreq.toPlainText().lower()
            mod_gl = 1
            found = False
            for i, mod in [('ms', 1000), ('us', 1000000), ('s', 1), ('m', 1/60), ('h', 1/3600)]:
                if frq != frq.replace(i, '') and not found:
                    frq = frq.replace(i, '')
                    mod_gl = mod
                    found = True
            frq = max(float(frq)/mod_gl, 0.0001) if frq.replace('.','').isnumeric() else 1
            print("Sample Period: "+str(frq)+"s")
            self.updateState('Create dataframe with T={}s sampling'.format(frq))
            df = self.MF4obj.dataFrame(self.selectedResults, frq)
            sz = df.shape
            self.updateState('Created a {} x {} dataframe (resampled with T={}s)'.format(sz[0], sz[1], frq))
            self.statusBar.setMaximum(1)
            from PySide6.QtWidgets import QFileDialog
            global default_path
            savepath = QFileDialog.getSaveFileName(self, 'Select a CSV to save the processed dataframe to.', default_path, 'Comma-separated values (*.csv)')[0]
            from pvevti.csvutil import df_to_csv
            from pvevti.genutil import parseFileSize
            self.statusBar.setMaximum(0)
            if savepath == '':
                savepath = default_path + "raw_export_"+str(1+len([pt for pt in os.listdir(default_path) if '.csv' in pt and 'raw_export_' in pt]))+".csv"
            df_to_csv(df, savepath, addition="", save_index=True)
            os.startfile(savepath)
            self.updateState('Saved to {} ({})'.format(savepath, parseFileSize(os.path.getsize(savepath))))
        except Exception as e:
            print("Error!")
            error = AllHasGoneAwry(self, "Error while processing the dataFrame.\nThis is most frequently caused by a malformed MF4 archive.\n  Exception: {}".format(str(e)), open=False, cancel=False)
            error.exec()
        self.statusBar.setMaximum(1)

    def updateSelectedList(self):
        from pvevti.mf4util import toName
        self.selectedList.clear()
        for item in self.selectedResults:
            self.selectedList.addItem(QListWidgetItem(toName(item)))
        if len(self.selectedResults) == 0:
            self.proceed.setEnabled(False)
            self.export_search.setVisible(False)
        else:
            self.proceed.setEnabled(True)
            self.export_search.setVisible(True)

    def changeSelected(self):
        idx = self.selectedList.currentRow()
        if idx != -1:
            self.selectedResults.pop(idx)
            self.updateSelectedList()
        
    def changeList(self):
        from pvevti.mf4util import toName
        idx = self.signalList.currentRow()
        print("ADD FEATURE")
        print(" Selected Index {}; Selected Value; {}".format(idx, self.searchResults[idx]))
        if self.searchResults[idx] not in self.selectedResults:
            self.selectedResults.append(self.searchResults[idx])
            print("Added",toName(self.searchResults[idx]),"to",[toName(x) for x in self.selectedResults])
        else:
            self.selectedResults.pop(self.selectedResults.index(self.searchResults[idx]))
        self.updateSelectedList()

    def clearAll(self):
        self.selectedResults = []
        self.updateSelectedList()
    
    def addAll(self):
        from pvevti.mf4util import toName
        print("ADD ALL FEATURES")
        for (idx,item) in enumerate([self.signalList.item(x) for x in range(self.signalList.count())]):
            if self.searchResults[idx] not in self.selectedResults and self.searchResults[idx]['Name'] != 't':
                self.selectedResults.append(self.searchResults[idx])
                print("Added",toName(self.searchResults[idx]),"to",[toName(x) for x in self.selectedResults])
        self.updateSelectedList()

    def addFromFile(self):
        # Get user input (GUI)
        txtFile = QFileDialog.getOpenFileName(self, 'Open a plain text file with comma-separated names.', default_path, 'Plain-text files (*.txt)')[0]

        # Accept valid text files
        if not os.path.isfile(txtFile) or '.txt' not in txtFile.lower():
            print("Invalid!")
            return
        
        # Read lines of text file
        lines = []
        with open(txtFile, 'r') as file:
            lines += [item.split(',') for item in file.readlines()]
        lines = [item.strip() for sublist in lines for item in sublist if item.strip() != '']
        
        # Display status to user
        showlines = ""
        if len(lines) > 0: showlines = ', '.join(lines[0:min(len(lines), 5)])
        if len(lines) > 5: showlines += ', ...'
        self.updateState('Trying to import signal(s): '+showlines)

        # Now to try to find signals!
        self.search_signals = lines
        del(lines)

        self.loadFileSignals()
    
    def updateAddAllBtn(self):
        if self.searchResults != []:
            self.add_all.setVisible(True)
        else:
            self.add_all.setVisible(False)
    
    def exportSearch(self):
        from PySide6.QtWidgets import QInputDialog
        # print(self.selectedResults)
        exportItems = [item['Name'] for item in self.selectedResults]
        print("Export!\n Items: {}".format(exportItems))
        exportStr = ",".join(exportItems)
        print(" String: {}".format(exportStr))
        listName, ok = QInputDialog.getText(self, 'Input Dialog', 'Enter the name of the signal list to create:')
        if not ok: listName = 'USER_SEARCHLIST'
        if '.txt' not in listName: listName += '.txt'
        for disallowed in [(' ', '_'), (',', '-'), '!', (':', '-'), '@', '#', '$', '%', '^', '&', '*', '(', ')', '=', '+', '?', ';', '"', "'", '`', ('\\', '-'), ('/', '-')]:
            if isinstance(disallowed, tuple):
                listName = listName.replace(disallowed[0], disallowed[1])
            else:
                listName = listName.replace(disallowed, '')
        with open(default_path+listName, 'w') as file:
            file.write(exportStr)
        self.updateState('Saved selected searchlist to '+default_path+listName)
        os.startfile(default_path)


    def loadFileSignals(self):
        from pvevti.mf4util import toName
        if not self.search_signals or not self.MF4obj or len(self.search_signals) == 0:
            print("Tried to load from cached signals. Failed.")
            return
        
        all_channels = self.MF4obj.channel_data
        search_results = [ch for ch in all_channels if any(signal == ch['Name'] for signal in self.search_signals)]
        for item in search_results:
            print(" Checking {} against current search results:".format(item))
            if item not in self.selectedResults:
                self.selectedResults.append(item)
                print("  Added {} to {}-item list".format(toName(item), len(self.selectedResults)))
        
        # Update
        print("Updating selected list")
        self.updateSelectedList()



    def updateListFromUnit(self):
        """
        Updates the list of channel search results based on a unit input.
        /deg is replaced by ° in the search algorithm. 
        """
        self.statusBar.setMaximum(0)
        app.processEvents()
        unitSearch = self.unitSearch.toPlainText()

        if unitSearch != '':
            for item in [('/deg', '°')]:
                print("Check {} against {}".format(item[0], unitSearch))
                if item[0] in unitSearch.lower():
                    print(" -> Check pass, replace")
                    unitSearch = unitSearch.replace(item[0], item[1])
            print(unitSearch)
            self.nameSearch.clear()
            self.sourceSearch.clear()
            all_channels = self.MF4obj.channel_data
            if len(unitSearch) >= 1:
                self.searchResults = [ch for ch in all_channels if unitSearch.lower() in ch['Unit'].lower()]
            else:
                self.searchResults = [ch for ch in all_channels if unitSearch.lower() == ch['Unit'].lower()]
            from pvevti.mf4util import toName
            newList = []
            for ch in self.searchResults:
                if ch not in newList:
                    newList.append(ch)
            self.searchResults = newList
            del newList
            self.updateState('Matched {} channel(s) by units'.format(len(self.searchResults)))
            self.signalList.clear()
            for ch in self.searchResults:
                newWidget = QListWidgetItem(toName(ch))
                newWidget.setToolTip("Group "+str(ch['Group'])+", source(s): "+ch['GroupSource'].replace(';', ','))
                self.signalList.addItem(newWidget)
        self.statusBar.setMaximum(1)
        app.processEvents()
        self.updateAddAllBtn()
    
    def updateListFromName(self):
        """
        Updates the list of channel search results based on a name input.
        If the user provides an '=' at the start of the query, only exact matches are returned.
        """
        self.statusBar.setMaximum(0)
        app.processEvents()
        nameSearch = self.nameSearch.toPlainText()
        if nameSearch != '' and self.MF4obj:
            print(nameSearch)
            self.unitSearch.clear()
            self.sourceSearch.clear()
            all_channels = self.MF4obj.channel_data
            if len(nameSearch) > 1 and nameSearch[0] == '=':
                self.searchResults = [ch for ch in all_channels if nameSearch[1:].lower() == ch['Name'].lower()]
            elif len(nameSearch) >= 3:
                # self.searchResults = [ch for ch in all_channels if nameSearch.lower() in ch['Name'].lower()]
                self.searchResults = []
                self.searchResults += [item for item in all_channels if nameSearch.lower() in item['Name'].lower()]
                self.searchResults = self.MF4obj.channels_by_name([ch['Name'] for ch in self.searchResults])
                print(" > Found {} ({})".format(len(self.searchResults), self.searchResults))
            else:
                self.searchResults = [ch for ch in all_channels if nameSearch.lower() == ch['Name'].lower()]
            from pvevti.mf4util import toName
            newList = []
            for ch in self.searchResults:
                if ch not in newList:
                    newList.append(ch)
            self.searchResults = newList
            del newList
            self.updateState('Matched {} channel(s) by name'.format(len(self.searchResults)))
            self.signalList.clear()
            for ch in self.searchResults:
                newWidget = QListWidgetItem(toName(ch))
                newWidget.setToolTip("Group "+str(ch['Group'])+", source(s): "+ch['GroupSource'].replace(';', ','))
                self.signalList.addItem(newWidget)
                
            #print("\n\nOverview of signals add:\n 'Matching': {}\n 'searchResults': {}".format(matching, self.searchResults))
        self.statusBar.setMaximum(1)
        app.processEvents()
        self.updateAddAllBtn()
    
    def updateListFromSource(self):
        """
        Updates the list of channel search results based on a source input.
        """
        self.statusBar.setMaximum(0)
        app.processEvents()
        sourceSearch = self.sourceSearch.toPlainText()
        if sourceSearch != '' and self.MF4obj:
            print(sourceSearch)
            self.unitSearch.clear()
            self.nameSearch.clear()
            all_channels = self.MF4obj.channel_data
            if len(sourceSearch) >= 3:
                # self.searchResults = [ch for ch in all_channels if nameSearch.lower() in ch['Name'].lower()]
                self.searchResults = []
                self.searchResults += [item for item in all_channels if sourceSearch.lower() in item['GroupSource'].lower()]
                print(" > Found {} ({})".format(len(self.searchResults), self.searchResults))
            else:
                self.searchResults = [ch for ch in all_channels if sourceSearch.lower() == ch['GroupSource'].lower()]
            from pvevti.mf4util import toName
            newList = []
            for ch in self.searchResults:
                if ch not in newList:
                    newList.append(ch)
            self.searchResults = newList
            del newList
            self.updateState('Matched {} channel(s) by source'.format(len(self.searchResults)))
            self.signalList.clear()
            for ch in self.searchResults:
                newWidget = QListWidgetItem(toName(ch))
                newWidget.setToolTip("Group "+str(ch['Group'])+", source(s): "+ch['GroupSource'].replace(';', ','))
                self.signalList.addItem(newWidget)
                
            #print("\n\nOverview of signals add:\n 'Matching': {}\n 'searchResults': {}".format(matching, self.searchResults))
        self.statusBar.setMaximum(1)
        app.processEvents()
        self.updateAddAllBtn()

    def check_open(self):
        """
        On text changed method, checks if the file is a valid MF4 and opens if it is.
        """
        txt = self.text_entry.toPlainText()
        if txt != '' and os.path.isfile(txt) and '.mf4' in txt.lower():
            self.updateState(' Opening file: {}'.format(txt))
            self.statusBar.setMaximum(0)
            app.processEvents()
            self.open_mf4()

            results_to_map = self.selectedResults
            self.selectedList.clear()
            self.clearAll()
            mapped_results = []
            if self.MF4obj:
                for old_item in results_to_map:
                    for new_item in self.MF4obj.channel_data:
                        if old_item['Name'] == new_item['Name'] and old_item['GroupSource'] == new_item['GroupSource']:
                            mapped_results.append(new_item)
                            print("Added {} to new list!".format(new_item))
                
                from pvevti.mf4util import toName
                for item in mapped_results:
                    self.selectedList.addItem(QListWidgetItem(toName(item)))
                    print("Added inferred {} to new selected list!".format(toName(item)))
            
            if len(mapped_results) == 0:
                self.proceed.setEnabled(False)
                self.export_search.setVisible(False)
            else:
                self.proceed.setEnabled(True)
                self.export_search.setVisible(True)

            # Clear out existing variables
            self.unitSearch.clear()
            self.nameSearch.clear()
            self.sourceSearch.clear()
            self.signalList.clear()
            self.searchResults =   []
            self.selectedResults = mapped_results
            self.add_all.setVisible(False)

            # Load in file signals
            self.loadFileSignals()
            self.statusBar.setMaximum(1)
            app.processEvents()

    def open_mf4(self):
        """
        Load MF4 object into the program.
        """
        from pvevti import mf4util
        print("Open "+self.text_entry.toPlainText())
        self.MF4obj = mf4util.MF4Object(self.text_entry.toPlainText())
        print(" Loaded "+str(self.MF4obj.num_channels)+" channels")
        self.updateState(' Loaded {} channels'.format(str(self.MF4obj.num_channels)))

    def pick_file(self):
        """
        Pick file method for GUI interaction when selecting a MF4 file. No return. Overwrites the text entry box.
        """
        from PySide6.QtWidgets import QFileDialog
        file = QFileDialog.getOpenFileName(self, 'Select the ASAM MDF file to import', default_path, "Measurement Data Format (*.mf4)")[0]
        print(file)
        if file != '':
            self.text_entry.setPlainText(file)
            self.mf4_path = file

# SLC Merge Dialog
class SLCPreferences(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QLabel, QPushButton, QVBoxLayout, QLineEdit
        layout = QVBoxLayout()
        label = QLabel("Enter a date or leave blank for today's date.")
        label.setFixedWidth(280)
        label.setStyleSheet('text-align:center;')
        self.textBox = QLineEdit()
        self.textBox.setPlaceholderText('Enter a date in the MM/DD/YYYY format.')
        self.textBox.setFixedWidth(280)
        btn = QPushButton('Continue')
        btn.clicked.connect(self.config_day)
        self.warn_label = QLabel("Correct formatting to match MM/DD/YYYY format.")
        self.warn_label.setVisible(False)
        self.warn_label.setStyleSheet('text-align:center; color:red;')
        layout.addWidget(label)
        layout.addWidget(self.textBox)
        layout.addWidget(btn)
        layout.addWidget(self.warn_label)
        self.setWindowTitle('Configure SLC Date')
        self.setLayout(layout)
    
    def config_day(self):

        global SLC_Date

        def parseNumber(x):
            if not str.isnumeric(x):
                return "00"
            else:
                if int(x) < 10:
                    return "0"+str(int(x))
                else:
                    return str(int(x))

        date = self.textBox.text().strip().split('/')
        date = [parseNumber(item) for item in date]
        if len(date) == 2:
            SLC_Date = '2025/'+'/'.join(date)
            self.close()
        elif len(date) == 3:
            SLC_Date = date[2]+'/'+'/'.join(date[0:1])
            self.close()
        elif self.textBox.text() != '':
            self.warn_label.setVisible(True)
        else:
            SLC_Date = ""
            self.close()

# Merging Preferences PQR Dialog
class MergePreferences(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.activePosition = 0
        self.filePath = ""
        from PySide6.QtWidgets import QVBoxLayout, QLabel, QButtonGroup, QRadioButton, QHBoxLayout, QPushButton, QPlainTextEdit
        btnGroup = QButtonGroup()
        # self.setWindowFlags(Qt.FramelessWindowHint)
        self.noReduction = QRadioButton('No Filtering')
        self.noReduction.toggled.connect(self.setNoReductionActive)
        self.noReduction.setChecked(True)
        btnGroup.addButton(self.noReduction)
        self.reduceByFile= QRadioButton('Filter by File')
        self.reduceByFile.toggled.connect(self.setReduceByFileActive)
        btnGroup.addButton(self.reduceByFile)
        self.reduceDirectly=QRadioButton('Filter Directly')
        self.reduceDirectly.toggled.connect(self.setReduceDirectlyActive)
        btnGroup.addButton(self.reduceDirectly)

        proceedButton = QPushButton('Proceed')
        proceedButton.clicked.connect(self.proceed)
        
        noReductionLayout = QVBoxLayout()
        noReductionLabel = QLabel('No parameters.')
        noReductionLayout.addStretch(1)
        noReductionLayout.addWidget(noReductionLabel)
        noReductionLayout.addStretch(1)

        reduceByFileLayout = QVBoxLayout()
        reduceByFileLabel = QLabel('Select the desired file to filter from.')
        reduceByFileSelect = QPushButton('Pick File')
        reduceByFileSelect.clicked.connect(self.pickPath)
        self.reduceByFileError = QLabel('Please select a file.')
        self.reduceByFileError.setStyleSheet('color:red;')
        self.reduceByFileError.setVisible(False)
        reduceByFileLayout.addStretch(1)
        reduceByFileLayout.addWidget(self.reduceByFileError)
        reduceByFileLayout.addWidget(reduceByFileLabel)
        reduceByFileLayout.addWidget(reduceByFileSelect)
        reduceByFileLayout.addStretch(1)

        reduceDirectlyLayout = QVBoxLayout()
        reduceDirectlyLabel = QLabel("Enter the signals below, separated with a ','")
        self.reduceDirectlyTextEntry = QPlainTextEdit()
        self.reduceDirectlyTextEntry.setPlaceholderText('Enter signals separated by comma')
        self.reduceDirectlyTextEntry.setFixedWidth(360)
        reduceDirectlyLayout.addStretch(1)
        reduceDirectlyLayout.addWidget(reduceDirectlyLabel)
        reduceDirectlyLayout.addWidget(self.reduceDirectlyTextEntry)
        reduceDirectlyLayout.addStretch(1)

        self.noReductionContainer = QWidget()
        self.reduceByFileContainer = QWidget()
        self.reduceDirectlyContainer = QWidget()
        self.noReductionContainer.setFixedWidth(360)
        self.noReductionContainer.setLayout(noReductionLayout)
        self.reduceByFileContainer.setFixedWidth(360)
        self.reduceByFileContainer.setLayout(reduceByFileLayout)
        self.reduceDirectlyContainer.setFixedWidth(360)
        self.reduceDirectlyContainer.setLayout(reduceDirectlyLayout)
        self.reduceByFileContainer.setVisible(False)
        self.reduceDirectlyContainer.setVisible(False)

        wrapper = QHBoxLayout()
        layoutWrapper = QWidget()
        layoutWrapper.setFixedWidth(120)
        layout = QVBoxLayout()
        layout.addWidget(self.noReduction)
        layout.addWidget(self.reduceByFile)
        layout.addWidget(self.reduceDirectly)
        layout.addWidget(proceedButton)
        layoutWrapper.setLayout(layout)
        wrapper.addWidget(layoutWrapper)
        wrapper.addWidget(self.noReductionContainer)
        wrapper.addWidget(self.reduceByFileContainer)
        wrapper.addWidget(self.reduceDirectlyContainer)
        self.setLayout(wrapper)
        self.setWindowTitle("Options for Merge")
    
    def setNoReductionActive(self, selected):
        if selected:
            print("No reduction active")
            self.reduceByFileContainer.setVisible(False)
            self.reduceDirectlyContainer.setVisible(False)
            self.noReductionContainer.setVisible(True)
            self.activePosition = 0
    def setReduceByFileActive(self, selected):
        if selected:
            print("Reduce by file")
            self.noReductionContainer.setVisible(False)
            self.reduceDirectlyContainer.setVisible(False)
            self.reduceByFileContainer.setVisible(True)
            self.reduceByFileError.setVisible(False)
            self.activePosition = 1
    def setReduceDirectlyActive(self, selected):
        if selected:
            print("Reduce active")
            self.noReductionContainer.setVisible(False)
            self.reduceByFileContainer.setVisible(False)
            self.reduceDirectlyContainer.setVisible(True)
            self.activePosition = 2

    def pickPath(self):
        self.filePath = QFileDialog.getOpenFileName(self, 'Select a file to source target signal names.', 
                                                    default_path, 'Text Files (*.txt)')[0]
        print("Path: "+self.filePath)
    def proceed(self):
        global filterSignals, combine_prefs
        if self.activePosition == 0:
            # State 0, no filtered signals. (All passed)
            filterSignals = []
            self.close()
        elif self.activePosition == 1:
            # State 1, reduction by means of file read
            print("Reduce by file!")
            if self.filePath == '' and combine_prefs == default_path:
                self.reduceByFileError.setVisible(True)
            elif combine_prefs == default_path:
                # Set default path, set path, open file, and read lines
                configure_default_path("COMBINEPREFS", self.filePath)
                combine_prefs = self.filePath
            with open(combine_prefs, 'r') as file:
                lines = file.readlines()

                # Loop through each line and extract the actual signal
                filterSignals = []
                for line in lines:
                    items = [item.strip().lower() for item in line.split(',') if item.strip() != '']
                    filterSignals += items
                print("Filtered items imported: "+str(filterSignals))
                self.close()
                pass # Do file mods
        else:
            # State 2, reduction by means of direct input
            rawInput = self.reduceDirectlyTextEntry.toPlainText()
            trimmedInput = [item.lower().strip() for item in rawInput.split(',') if item.strip() != '']
            filterSignals = trimmedInput
            self.close()

# Incorrect path dialog
class AllHasGoneAwry(QDialog):
    def __init__(self, parent=None, label="The path for the runlog script is not correct.\nOpen the correct file.", open = True, cancel = True):
        super().__init__(parent)
        from PySide6.QtWidgets import QVBoxLayout, QLabel
        self.setWindowTitle("Catastrophe!")
        if open and cancel:
            QBtn = (QDialogButtonBox.Open | QDialogButtonBox.Cancel)
        elif open:
            QBtn = QDialogButtonBox.Open
        elif cancel:
            QBtn = QDialogButtonBox.Cancel
        else:
            QBtn = QDialogButtonBox.Ok

        self.buttonBox = QDialogButtonBox(QBtn)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)

        layout = QVBoxLayout()
        message = QLabel(label)
        layout.addWidget(message)
        layout.addWidget(self.buttonBox)
        self.setLayout(layout)

# Main Window
class MainWindow(QtWidgets.QMainWindow):
    from pvevti import pdfutil, csvutil, genutil

    def __init__(self):
        from PySide6.QtWidgets import (QComboBox, QCheckBox, QSizePolicy,
                               QVBoxLayout, QWidget, QLabel, QHBoxLayout, 
                               QToolBar, QPushButton, QListWidget, QListWidgetItem)
        from PySide6.QtGui import QAction, QIcon
        super().__init__()

        self.setWindowTitle("DTNA/PVE File Processing Tool")
        
        l1 = QVBoxLayout()
        l1.setContentsMargins(0,0,0,0)
        l1.setSpacing(0) # Spacing between boxes in layout

        titleLabel = QLabel('File Processing Tool')
        titleLabel.setStyleSheet(titleStyle)
        titleLabel.setMargin(0)
        titleLabel.setFixedHeight(28)
        subtitle = QLabel('pvevti '+pvevti_version+', FPT '+FPT_version+', PySide6 '+PySide6_version)
        subtitle.setStyleSheet(subtitleStyle)
        subtitle.setMargin(0)
        subtitle.setFixedHeight(19)
        l1.addWidget(titleLabel)
        l1.addWidget(subtitle)

        l2 = QHBoxLayout()
        leftPane = QWidget()
        leftPane.setStyleSheet(mainPane)
        leftPane.setFixedWidth(240)

        # Top level left pane
        prefsLayout = QVBoxLayout()
        prefsImportTitle  = QLabel('Import')
        prefsImportTitle.setStyleSheet(prefsTitleStyle)
        prefsProcessingTitle  = QLabel('Processing')
        prefsProcessingTitle.setStyleSheet(prefsTitleStyle)
        prefsExportTitle  = QLabel('Export')
        prefsExportTitle.setStyleSheet(prefsTitleStyle)

        
        ## Imports

        # Encoding Setup
        encodingLayout = QHBoxLayout()
        self.encoding = QComboBox()
        self.encoding.setAccessibleName('Encoding')
        self.encoding.setStyleSheet(smallText)
        self.encoding.setMaximumWidth(70)
        encodeTypes = ['ascii', 'latin-1', 'utf-8', 'utf-16']
        for encodeType in encodeTypes:
            self.encoding.addItem(encodeType)
        self.encoding.setCurrentIndex(encodeTypes.index('latin-1'))
        encodingLabel = QLabel('Encoding')
        encodingLabel.setStyleSheet(smallTextIndent)
        encodingLayout.addWidget(encodingLabel)
        encodingLayout.addWidget(self.encoding)
        
        # Memory Reservations
        memoryLayout = QHBoxLayout()
        self.memory = QComboBox()
        self.memory.setAccessibleName('Memory')
        self.memory.setStyleSheet(smallText)
        self.memory.setMaximumWidth(70)
        memoryReservations = ['low', 'normal']
        for memoryReservation in memoryReservations:
            self.memory.addItem(memoryReservation)
        self.memory.setCurrentIndex(memoryReservations.index('normal'))
        memoryLabel = QLabel('Memory Reservation')
        memoryLabel.setStyleSheet(smallTextIndent)
        memoryLayout.addWidget(memoryLabel)
        memoryLayout.addWidget(self.memory)
        
        # Processing
        processingLayout = QVBoxLayout()
        processingLabel = QLabel("Processing:")
        processingLabel.setStyleSheet(smallIndent)
        processingLayout.addWidget(processingLabel)
        
        # GPS
        gpsLayout = QHBoxLayout()
        self.gps = QComboBox()
        self.gps.setAccessibleName('GPS Filter')
        self.gps.setStyleSheet(smallText)
        self.gps.setMaximumWidth(70)
        gpsOptions = ['Enabled', 'Disabled']
        for gpsOption in gpsOptions:
            self.gps.addItem(gpsOption)
        self.gps.setCurrentIndex(gpsOptions.index('Enabled'))
        gpsLabel = QLabel('GPS Filter')
        gpsLabel.setStyleSheet(smallTextIndent)
        gpsLayout.addWidget(gpsLabel)
        gpsLayout.addWidget(self.gps)
        processingLayout.addLayout(gpsLayout)
        
        # Compression
        compLayout = QHBoxLayout()
        self.compression = QComboBox()
        self.compression.setAccessibleName('Compression')
        self.compression.setStyleSheet(smallText)
        self.compression.setMaximumWidth(70)
        compressionOptions = ['Aggressive', 'Normal', 'Reduced', 'Disabled']
        for compressionOption in compressionOptions:
            self.compression.addItem(compressionOption)
        self.compression.setCurrentIndex(compressionOptions.index('Normal'))
        compressionLabel = QLabel('Compression')
        compressionLabel.setStyleSheet(smallTextIndent)
        compLayout.addWidget(compressionLabel)
        compLayout.addWidget(self.compression)
        processingLayout.addLayout(compLayout)
        

        ## Reports
        reportLayout = QVBoxLayout()
        reportLabel = QLabel("Generate Reports:")
        reportLabel.setStyleSheet(smallIndent)
        reportLayout.addWidget(reportLabel)
        global boxes, reports
        boxes = []
        for report in reports:
            checkBox = QCheckBox()
            checkBox.setAccessibleName(report)
            checkBox.setStyleSheet(checkBoxStyle)
            label = QLabel(report)
            label.setStyleSheet(smallTextIndent)
            layout = QHBoxLayout()
            layout.addWidget(label)
            layout.addStretch(1)
            layout.addWidget(checkBox)
            boxes.append(checkBox)
            reportLayout.addLayout(layout)


        ## Export
        exportLayout = QVBoxLayout()

        # Label
        doExportLabel = QLabel('Export Processed CSVs')
        doExportLabel.setStyleSheet(smallTextIndent)

        # Toggle
        self.doExport = QCheckBox()
        self.doExport.setAccessibleName('Export Files')
        self.doExport.setChecked(True)
        self.doExport.setStyleSheet(checkBoxStyle)
        self.doExport.clicked.connect(self.toggleDoExport)

        # Layout
        doExportLayout = QHBoxLayout()
        doExportLayout.addWidget(doExportLabel)
        doExportLayout.addStretch(1)
        doExportLayout.addWidget(self.doExport)

        replaceOriginalsLayoutContainer = QWidget()

        # Label
        replaceOriginalsLabel = QLabel('Replace Originals')
        replaceOriginalsLabel.setStyleSheet(smallTextIndent)

        # Toggle
        self.replaceOriginals = QCheckBox()
        self.replaceOriginals.setAccessibleName('Replace Originals')
        self.replaceOriginals.setChecked(False)
        self.replaceOriginals.setStyleSheet(checkBoxStyle)
        self.replaceOriginals.clicked.connect(self.toggleReplaceOriginals)

        # Layout
        replaceOriginalsLayout = QHBoxLayout(replaceOriginalsLayoutContainer)
        replaceOriginalsLayout.addWidget(replaceOriginalsLabel)
        replaceOriginalsLayout.addStretch(1)
        replaceOriginalsLayout.addWidget(self.replaceOriginals)
        replaceOriginalsLayout.setContentsMargins(0,0,0,0)
        self.doExport.toggled.connect(replaceOriginalsLayoutContainer.setVisible)

        exportLayout.addLayout(doExportLayout)
        exportLayout.addWidget(replaceOriginalsLayoutContainer)


        ## Save directories
        csvSaveDirLayoutContainer = QWidget()
        csvSaveDirLayout = QHBoxLayout(csvSaveDirLayoutContainer)
        csvSaveDirLayout.setContentsMargins(0,0,0,0)
        csvSaveDirLabel = QLabel('CSV Export Directory')
        csvSaveDirLabel.setStyleSheet(smallTextIndent)
        csvSaveDirBtn = QPushButton('Select')
        csvSaveDirBtn.setStyleSheet(buttonStyle)
        csvSaveDirBtn.clicked.connect(self.pickCSVSaveDir)
        csvSaveDirLayout.addWidget(csvSaveDirLabel)
        csvSaveDirLayout.addStretch(1)
        csvSaveDirLayout.addWidget(csvSaveDirBtn)

        repSaveDirLayoutContainer = QWidget()
        repSaveDirLayout = QHBoxLayout(repSaveDirLayoutContainer)
        repSaveDirLayout.setContentsMargins(0,0,0,0)
        repSaveDirLabel = QLabel('Report Export Directory')
        repSaveDirLabel.setStyleSheet(smallTextIndent)
        repSaveDirBtn = QPushButton('Select')
        repSaveDirBtn.setStyleSheet(buttonStyle)
        repSaveDirBtn.clicked.connect(self.pickReportSaveDir)
        repSaveDirLayout.addWidget(repSaveDirLabel)
        repSaveDirLayout.addStretch(1)
        repSaveDirLayout.addWidget(repSaveDirBtn)

        # Add preferences widgets
        prefsLayout.addWidget(prefsImportTitle)
        prefsLayout.addLayout(encodingLayout)
        prefsLayout.addLayout(memoryLayout)
        prefsLayout.addWidget(prefsProcessingTitle)
        prefsLayout.addLayout(processingLayout)
        prefsLayout.addLayout(reportLayout)
        prefsLayout.addWidget(prefsExportTitle)
        prefsLayout.addLayout(exportLayout)
        prefsLayout.addSpacing(10)
        prefsLayout.addWidget(csvSaveDirLayoutContainer)
        prefsLayout.addWidget(repSaveDirLayoutContainer)
        prefsLayout.addStretch(1)

        # Child the prefs layout to the left pane
        leftPane.setLayout(prefsLayout)
        l2.addWidget(leftPane)
        
        # Create placeholder right pane for previews
        rightPane = QWidget()
        rightPane.setStyleSheet(mainPane)
        rightPane.setMinimumWidth(720)

        viewLayout = QVBoxLayout()
        viewLayout.setContentsMargins(0,0,0,0)
        viewLayout.setSpacing(0)

        buttonParent = QWidget()
        buttonContainer = QHBoxLayout()

        addFileBtn = QPushButton(icon=QIcon(resource_path("icons/file-plus.svg")))
        addFileBtn.setStyleSheet(iconButtonStyle)
        addFileBtn.clicked.connect(self.pickFile)
        addFileBtn.setToolTip('Open CSV file(s)\nShortcut: Ctrl+O')
        addFileBtnSC = QShortcut(QKeySequence('Ctrl+O'), self)
        addFileBtnSC.activated.connect(self.pickFile)

        pasteSC = QShortcut(QKeySequence.Paste, self)
        pasteSC.activated.connect(self.pasteFiles)
        
        addPathBtn = QPushButton(icon=QIcon(resource_path("icons/folder-plus.svg")))
        addPathBtn.setStyleSheet(iconButtonStyle)
        addPathBtn.clicked.connect(self.pickFilesInPath)
        addPathBtn.setToolTip('Load all CSV files in path\nShortcut: Ctrl+Shift+O')
        addPathBtnSC = QShortcut(QKeySequence('Ctrl+Shift+O'), self)
        addPathBtnSC.activated.connect(self.pickFilesInPath)
        
        addSLCPathBtn = QPushButton(icon=QIcon(resource_path("icons/journal-plus.svg")))
        addSLCPathBtn.setStyleSheet(iconButtonStyle)
        addSLCPathBtn.clicked.connect(self.pickSLCFilesInPath)
        addSLCPathBtn.setToolTip('Load all SLC files in path\nShortcut: Ctrl+Alt+O')
        addSLCSC = QShortcut(QKeySequence('Ctrl+Alt+O'), self)
        addSLCSC.activated.connect(self.pickSLCFilesInPath)

        trashBtn = QPushButton(icon=QIcon(resource_path("icons/folder-x.svg")))
        trashBtn.setStyleSheet(iconButtonStyle)
        trashBtn.clicked.connect(self.clearItems)
        trashBtn.setToolTip('Discard all loaded CSV files\nShortcut: Ctrl+Delete')
        trashFilesSC = QShortcut(QKeySequence('Ctrl+Delete'), self)
        trashFilesSC.activated.connect(self.clearItems)

        trashFileBtn = QPushButton(icon=QIcon(resource_path("icons/file-x.svg")))
        trashFileBtn.setStyleSheet(iconButtonStyle)
        trashFileBtn.clicked.connect(self.clearItem)
        trashFileBtn.setToolTip('Discard selected CSV file\nShortcut: Delete')
        trashFileSC = QShortcut(QKeySequence.Delete, self)
        trashFileSC.activated.connect(self.clearItem)

        metaReport = QPushButton(icon=QIcon(resource_path("icons/geo.svg")))
        metaReport.setStyleSheet(iconButtonStyle)
        metaReport.clicked.connect(self.createMetaReport)
        metaReport.setToolTip('Create a high level document with route information\nShortcut: Ctrl+Shift+H')
        metaReportSC = QShortcut(QKeySequence('Ctrl+Shift+H'), self)
        metaReportSC.activated.connect(self.createMetaReport)

        quickCSV = QPushButton(icon=QIcon(resource_path("icons/filetype-csv.svg")))
        quickCSV.setStyleSheet(iconButtonStyle)
        quickCSV.clicked.connect(self.processFiles)
        quickCSV.setToolTip('Quick-create Processed CSVs\nShortcut: Ctrl+Shift+P')
        quickCSVSC = QShortcut(QKeySequence('Ctrl+Shift+P'), self)
        quickCSVSC.activated.connect(self.processFiles)

        mergeCSV = QPushButton(icon=QIcon(resource_path("icons/intersect.svg")))
        mergeCSV.setStyleSheet(iconButtonStyle)
        mergeCSV.clicked.connect(self.mergeFiles)
        mergeCSV.setToolTip('Merge Selected CSVs\nHold shift to open options\nShortcut: Ctrl+Shift+M')
        mergeCSVSC = QShortcut(QKeySequence('Ctrl+Shift+M'), self)
        mergeCSVSC.activated.connect(self.mergeFiles)
        
        quickPDF = QPushButton(icon=QIcon(resource_path("icons/file-richtext.svg")))
        quickPDF.setStyleSheet(iconButtonStyle)
        quickPDF.clicked.connect(self.makeReports)
        quickPDF.setToolTip('Quick-create PDF Reports\nShortcut: Ctrl+Shift+R')
        quickPDFSC = QShortcut(QKeySequence('Ctrl+Shift+R'), self)
        quickPDFSC.activated.connect(self.makeReports)

        runBtn = QPushButton(icon=QIcon(resource_path("icons/gear.svg")))
        runBtn.setStyleSheet(iconButtonStyle)
        runBtn.clicked.connect(self.processAndReport)
        runBtn.setToolTip('Execute program according to all preferences')

        buttonContainer.setContentsMargins(0,0,0,0)
        buttonContainer.setSpacing(0)
        buttonContainer.addSpacing(2)
        buttonContainer.addWidget(addFileBtn)
        buttonContainer.addWidget(addPathBtn)
        buttonContainer.addWidget(addSLCPathBtn)
        buttonContainer.addSpacing(30)
        buttonContainer.addWidget(trashBtn)
        buttonContainer.addWidget(trashFileBtn)
        buttonContainer.addStretch(1)
        buttonContainer.addWidget(metaReport)
        buttonContainer.addWidget(mergeCSV)
        buttonContainer.addWidget(quickCSV)
        buttonContainer.addWidget(quickPDF)
        buttonContainer.addSpacing(30)
        buttonContainer.addWidget(runBtn)
        buttonParent.setLayout(buttonContainer)
        buttonParent.setContentsMargins(0,0,0,0)
        buttonParent.setStyleSheet(buttonParentStyle)

        viewLayout.addWidget(buttonParent)
        quickView = QShortcut(QKeySequence('Space'), self)
        quickView.activated.connect(self.quickView)

        self.progressBar = QProgressBar()
        self.progressBar.setValue(50)
        self.progressBar.setVisible(False)
        self.progressBar.setTextVisible(False)
        self.progressBar.setStyleSheet(progressBarStyle)
        self.progressBar.setFixedHeight(3)
        viewLayout.addWidget(self.progressBar)
        
        viewLayout.addSpacing(2)

        self.filesView = QListWidget()
        self.filesView.setStyleSheet(listwidgetStyle)
        items = ["Load CSV Files to get started."]
        for item in items:
            self.filesView.addItem(QListWidgetItem(item))
        viewLayout.addWidget(self.filesView)
        
        viewLayout.addSpacing(2)
        self.statusLabel = QLabel('')
        self.statusLabel.setStyleSheet(statusLabelStyle)
        self.savedCSV = QPushButton('View Saved CSVs')
        self.savedCSV.clicked.connect(self.openCSVDir)
        self.savedCSV.setVisible(False)
        self.savedCSV.setStyleSheet(viewBtnStyle)
        self.savedRep = QPushButton('View Saved Reports')
        self.savedRep.clicked.connect(self.openReportDir)
        self.savedRep.setVisible(False)
        self.savedRep.setStyleSheet(viewBtnStyle)
        statusContainer = QHBoxLayout()
        statusContainer.addWidget(self.statusLabel)
        statusContainer.addStretch(1)
        statusContainer.addWidget(self.savedCSV)
        statusContainer.addWidget(self.savedRep)
        viewLayout.addLayout(statusContainer)

        viewLayout.addSpacing(1)

        rightPane.setLayout(viewLayout)
        l2.addWidget(rightPane)
        l2.addSpacing(3)

        # Add l2 layout to the l1 stack
        l1.addLayout(l2)

        # Create an empty widget, fill with layout, and assign.
        widget = QWidget()
        widget.setLayout(l1)
        widget.setStyleSheet('background-color:#e6e6e6;')
        self.setCentralWidget(widget)

        # Build out toolbar
        toolbar = QToolBar('Toolbar')
        toolbar.toggleViewAction().setEnabled(False)
        toolbar.setStyleSheet(toolbarButtonStyle)

        # Pick single/multi file
        pickfile_button = QAction("Insert from clipboard", self)
        pickfile_button.setToolTip('Enter files on the clipboard into the workspace')
        pickfile_button.triggered.connect(self.pasteFiles)

        # Automatically scan for new CSVs
        autoscan_button = QAction("Autoscan", self)
        autoscan_button.setToolTip('Autoscan files from the last opened directory')
        autoscan_button.triggered.connect(self.autoscan)

        # Update runlog
        runlog_button = QAction("Update runlog", self)
        runlog_button.setToolTip('Update the runlog file')
        runlog_button.triggered.connect(self.execute_runlog)
        
        # Open runlog file
        runlog_open_button = QAction("Open runlog", self)
        runlog_open_button.setToolTip('Open the runlog file')
        runlog_open_button.triggered.connect(self.open_runlog)

        # Configure PQR Merge
        merge_config_button = QAction("PQR Merge Parameters", self)
        merge_config_button.setToolTip('Select the PQR Merge signal parameters')
        merge_config_button.triggered.connect(self.config_merge)

        # Daily SLC
        daily_SLC_button = QAction("Create Daily SLC", self)
        daily_SLC_button.setToolTip('Create an SLC report from today\nHold shift to specify the date')
        daily_SLC_button.triggered.connect(self.SLC_Report)

        # MF4 Convert
        mf4_button = QAction("Open MF4 Tool", self)
        mf4_button.setToolTip('Start the MF4 processing interface')
        mf4_button.triggered.connect(self.mf4_process)

        # MF4 Convert
        custom_button = QAction("Custom Reports", self)
        custom_button.setToolTip('Open the custom reports interface and restarts the application\nHold shift to clear custom reports')
        custom_button.triggered.connect(self.import_custom_dialog)

        # Automatically scan for new CSVs
        close_button = QAction("Close FPT", self)
        close_button.setToolTip('Close the program')
        close_button.triggered.connect(self.closeApp)

        # Child buttons to toolbar
        buffer = QWidget()
        buffer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        toolbar.addAction(pickfile_button)
        toolbar.addAction(autoscan_button)
        toolbar.addAction(runlog_button)
        toolbar.addAction(runlog_open_button)
        toolbar.addAction(merge_config_button)
        toolbar.addAction(daily_SLC_button)
        toolbar.addAction(mf4_button)
        toolbar.addWidget(buffer)
        toolbar.addAction(custom_button)
        toolbar.addSeparator()
        toolbar.addAction(close_button)

        # Child toolbar to self
        self.addToolBar(toolbar)
        self.show()

    def mf4_process(self):
        MF4Dialog(self).exec()

    def SLC_Report(self):
        modifiers = QApplication.queryKeyboardModifiers()
        if modifiers & Qt.ShiftModifier:
            print("Pressed with shift held.")
            SLCPreferences(self).exec()

        from datetime import date
        global SLC_Date
        if SLC_Date == '':
            search_date = date.today().strftime("%Y_%m_%d")
        else:
            search_date = '_'.join(SLC_Date.split('/'))
        
        matching_files = []
        search_dir = QFileDialog.getExistingDirectory(self, 'Select the directory to recursively search.', csv_open_default_path)
        for root, dirs, files in os.walk(search_dir):
            for file in [item for item in files if 'SLC' in item and search_date in item]:
                matching_files.append(os.path.join(root, file).replace('\\', '/'))
        
        if len(matching_files) == 0:
            from PySide6.QtWidgets import QMessageBox
            notice = QMessageBox.information(self, "No inputs!", "There are no SLCs in the provided directory and subdirectories from "+search_date+".\nEnsure the date is correct and files exist if this is unexpected.")
        else:
            name = "DailyReport_"+search_date
            self.mergeFiles(matching_files, name)

    def import_custom_dialog(self):
        modifiers = QApplication.queryKeyboardModifiers()
        if modifiers & Qt.ShiftModifier:
            print("Pressed with shift held.")
            clear_custom_paths()
            self.restartApp()
            return
        
        # Otherwise, continue
        print(" > Import custom report")

        global custom_reports
        import json

        print("   > Query user")
        newPath = QFileDialog.getOpenFileName(self, 'Select a JSON File with the configured document layout.', csv_open_default_path, "Layout file (*.json)")[0]
        
        # If unique
        if newPath not in custom_reports and '.json' in newPath and os.path.isfile(newPath):
            with open(newPath, 'r') as file:
                data_dict = json.load(file)
            print(data_dict)
            reportName = "CUSTOMREPORT_"+data_dict['docTitle']
            custom_reports.append(newPath)
            configure_default_path(reportName, newPath)
        self.restartApp()

    def config_merge(self):
        print("Pulling preferences for merge")
        MergePreferences(self).exec()

    def updateProgress(self, step=0, possible=1):
        if step == 0 or step >= possible:
            self.progressBar.setVisible(False)
            self.progressBar.setValue(0)
            QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)
        else:
            self.progressBar.setVisible(True)
            self.progressBar.setValue(100*step/possible)
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)

        app.processEvents()


    ## File explorer utilization
    def openCSVDir(self):
        global savedir_csv
        if os.path.isdir(savedir_csv):
            os.startfile(savedir_csv)
        print(savedir_csv)

    def openReportDir(self):
        global savedir_rep
        if os.path.isdir(savedir_rep):
            os.startfile(savedir_rep)
        print(savedir_rep)


    ## UI updater method
    def updateState(self, txt=""):
        self.savedRep.setVisible(False)
        self.savedCSV.setVisible(False)
        if len(txt) > 85:
            self.statusLabel.setText(txt[0:65]+" ... "+txt[-20:])
        else:
            self.statusLabel.setText(txt)
        app.processEvents()


    ## Directory selection methods
    def pickCSVSaveDir(self):
        global savedir_csv
        savedir_csv = QFileDialog.getExistingDirectory(self, "Select a directory to save processed CSV files to.", csv_save_default_path)
        if savedir_csv != "":
            if savedir_csv[-1] == "/":
                configure_default_path("LASTCSVSAVEPATH", savedir_csv)
                self.updateState('CSV path set to '+savedir_csv)
            else:
                configure_default_path("LASTCSVSAVEPATH", savedir_csv+"/")
                self.updateState('CSV path set to '+savedir_csv+"/")

    def pickReportSaveDir(self):
        global savedir_rep
        savedir_rep = QFileDialog.getExistingDirectory(self, "Select a directory to save processed reports to.", rep_save_default_path)
        
        if savedir_rep != "":
            if savedir_rep[-1] == "/":
                configure_default_path("LASTREPORTSAVEPATH", savedir_rep)
                self.updateState('Report path set to '+savedir_rep)
            else:
                configure_default_path("LASTREPORTSAVEPATH", savedir_rep+"/")
                self.updateState('Report path set to '+savedir_rep+"/")
    

    ## Toggle methods to manage UI elements

    def toggleDoExport(self):
        global doExport
        doExport = not doExport
        print("Export: New state",doExport)

    def toggleReplaceOriginals(self):
        global replaceOriginals
        replaceOriginals = not replaceOriginals
        print("Replace: New state",replaceOriginals)
        
    
    ## Loading and unloading file paths into/from the list

    def pasteFiles(self):
        global paths
        rawText = QGuiApplication.clipboard().text()
        newpaths = rawText.split('\n')
        newpaths = [path.replace('file:///', '') for path in newpaths if 'file:///' in path]
        newpaths = [path for path in newpaths if os.path.isfile(path)]
        print("Paste", newpaths)
        len1 = len(paths)
        self.importFiles(newpaths)
        self.updateState('Pasted '+str(len(paths)-len1)+' CSV file(s).')
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)
    
    def importFiles(self, files):
        global paths
        from pvevti.genutil import parseFileSize
        if len(files) > 0:
            print("/".join(files[-1].split('/')[0:-1])+"/")
            configure_default_path("LASTCSVGETPATH", "/".join(files[-1].split('/')[0:-1])+"/")
        all_items = [self.filesView.item(i).text() for i in range(self.filesView.count())]
        if all_items == ['Load CSV Files to get started.'] and len(files) > 0:
            paths = []
            self.filesView.clear()
        num = 0
        for file in files:
            if file not in paths:
                num += 1
                print(" > "+file)
                paths.append(file)
                sz = os.path.getsize(file)
                self.filesView.addItem(QListWidgetItem("["+parseFileSize(sz)+"] "+file))
        self.updateState('Imported '+str(num)+' file(s)')
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)

    def pickFile(self):
        print(" == IMPORT ==")
        for item in [self.gps, self.compression, self.encoding, self.memory]:
            print(item.accessibleName(), ":", item.currentText())

        files = QFileDialog.getOpenFileNames(self, "Select one (or more) CSVs to open.", csv_open_default_path, "Comma-Separated Values (*.csv)")[0]
        self.importFiles(files)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)
    
    def pickFilesInPath(self):
        from PySide6.QtWidgets import QInputDialog
        modifiers = QApplication.queryKeyboardModifiers()
        if modifiers & Qt.ShiftModifier:
            print("Pressed with shift held.")
            filterText, ok = QInputDialog.getText(self, 'Add Filters (Space separated)', 'Please enter any filters below, separated by spaces.')
        else:
            print("Pressed.")
            filterText = ''
            ok = False

        if ok:
            print("Unparsed Filters: {}".format(filterText))
            filterText = [st.strip() for st in filterText.split(' ')]
            include = [st[1:] for st in filterText if len(st) > 1 and st[0] == '+']
            exclude = [st[1:] for st in filterText if len(st) > 1 and st[0] == '-']
            print("Parsed:\n  Include -> [{}]\n  Exclude -> [{}]".format("; ".join(include), "; ".join(exclude)))

        from pvevti.genutil import parseFileSize
        global paths
        print(" == IMPORT ==")
        for item in [self.gps, self.compression, self.encoding, self.memory]:
            print("  ",item.accessibleName(), ":", item.currentText())
        dir = QFileDialog.getExistingDirectory(self, "Select a directory to search")
        if dir != '':
            files = [dir+'/'+file for file in os.listdir(dir) if '.csv' in file]
            if ok:
                print(" > Unfiltered: {}".format(files))
                files = [file for file in files if include == [] or any(x in file for x in include)] # Included filter
                print(" > Filtered-1: {}".format(files))
                files = [file for file in files if exclude == [] or not any(x in file for x in exclude)] # Excluded filter
                print(" > Filtered-2: {}".format(files))
        else: 
            files = []
        if len(files) > 0:
            print(" > Check path: "+"/".join(files[-1].split('/')[0:-1])+"/")
            configure_default_path("LASTCSVGETPATH", "/".join(files[-1].split('/')[0:-1])+"/")
        all_items = [self.filesView.item(i).text() for i in range(self.filesView.count())]
        if all_items == ['Load CSV Files to get started.'] and len(files) > 0:
            paths = []
            self.filesView.clear()
        num = 0
        i = 0
        for file in files:
            i += 1
            if file not in paths:
                self.updateProgress(i, len(files))
                num += 1
                print(" > Add File: "+file)
                paths.append(file)
                sz = os.path.getsize(file)
                self.filesView.addItem(QListWidgetItem("["+parseFileSize(sz)+"] "+file))
        self.updateState('Imported '+str(num)+' file(s) from directory '+dir)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)

    def pickSLCFilesInPath(self):
        from pvevti.genutil import parseFileSize
        global paths
        print(" == IMPORT ==")
        for item in [self.gps, self.compression, self.encoding, self.memory]:
            print(item.accessibleName(), ":", item.currentText())
        dir = QFileDialog.getExistingDirectory(self, "Select a directory to search")
        if dir != '':
            files = [dir+'/'+file for file in os.listdir(dir) if '.csv' in file and 'SLC' in file]
        else: 
            files = []
        if len(files) > 0:
            print("/".join(files[-1].split('/')[0:-1])+"/")
            configure_default_path("LASTCSVGETPATH", "/".join(files[-1].split('/')[0:-1])+"/")
        all_items = [self.filesView.item(i).text() for i in range(self.filesView.count())]
        if all_items == ['Load CSV Files to get started.'] and len(files) > 0:
            paths = []
            self.filesView.clear()
        num = 0
        for file in files:
            if file not in paths:
                num += 1
                print(" > "+file)
                paths.append(file)
                sz = os.path.getsize(file)
                self.filesView.addItem(QListWidgetItem("["+parseFileSize(sz)+"] "+file))
        self.updateState('Imported '+str(num)+' file(s) from directory '+dir)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)

    def configure_runlog(self, config_file=True, config_dir=True):
        """
        Attempts to configure the runlog directory. If it fails or no file is provided by the user, returns -1. 
        Else, returns the filepath for the excel-type runlog file.
        """
        print("Configure Runlog")

        if config_dir:
            directory = QFileDialog.getExistingDirectory(self, 'Select the search directory.', default_path)
            print('>',directory)
            if directory != '':
                configure_default_path("RUNLOGDIR", directory)
                print("Configured runlog directory at "+directory)
                global runlog_dir
                runlog_dir = directory

        if config_file:
            file = QFileDialog.getOpenFileName(self, 'Select existing runlog file.', default_path, "Excel Files (*.xlsx)")
            print('>',file)
            if file != ('', ''):
                configure_default_path("RUNLOG", file[0])
                print("Configured runlog at "+file[0])
                global runlog
                runlog = file[0]
                return file[0]
        return -1
    
    def open_runlog(self):
        global runlog
        if '.xlsx' in runlog and os.path.exists(runlog):
            os.startfile(runlog)
        else:
            query = AllHasGoneAwry(self)
            if query.exec():
                self.configure_runlog(config_dir=False)
                self.open_runlog()

    def execute_runlog(self):
        modifiers = QApplication.queryKeyboardModifiers()
        if modifiers & Qt.ShiftModifier:
            print("Pressed with shift held.")
            self.configure_runlog()
        else:
            print("Pressed.")

        
        # Check the runlog path exists and has not been malformed
        global runlog
        global runlog_dir
        from datetime import datetime
        from pandas import DataFrame, read_excel

        # If the runlog is real, execute
        if '.xlsx' in runlog and os.path.exists(runlog):

            # Import the excel data and check which files already exist
            excelData = read_excel(runlog, usecols=['Data File Name'])
            existing_files = sum(excelData.values.tolist(), [])

            # Create empty filedata list, to later become a dataframe
            filedata = []
            print("Runlog exists! Executing.")

            # Lazy load imports
            from pandas import read_excel
            from openpyxl import load_workbook
            from pvevti.csvutil import df_from_csv
            from pvevti.genutil import getRoute

            # Walk through all directories and subdirectories
            for root, dirs, files in os.walk(runlog_dir):

                # Only look at PQR files with the .csv appendage
                for file in [file for file in files if "pqr" in file.lower() and ".csv" in file.lower()]:
                    path = os.path.join(root, file)

                    # If the file doesn't already exist in the runlog, execute GPS route detection
                    if file not in existing_files:
                        df = df_from_csv(path, ["GPS_x[°]", "GPS_y[°]"])
                        if "GPS_x[°]" in df.columns and "GPS_y[°]" in df.columns:
                            route = getRoute(df)[0]
                    else:
                        route = ""

                    # Fill the filedata list
                    filedata.append({
                        'File Name': file,
                        'File Size (KB)': round(os.path.getsize(path)/1000, 1),
                        'Last Modified Time': datetime.fromtimestamp(int(os.path.getmtime(path))),
                        'Route': route
                    })

            # Convert to a parsed dataframe where only the unshared values are kept
            filedata = DataFrame(filedata).dropna(how='all')
            filedata = filedata[~filedata['File Name'].isin(excelData['Data File Name'])]
            filedata.reset_index(drop=True, inplace=True)
            
            # Load the excel workbook and find the last row
            workbook = load_workbook(runlog)
            sheet = workbook.active
            last_row = sheet.max_row
            while last_row > 0 and all(sheet.cell(row=last_row, column=col).value is None for col in range(1, sheet.max_column + 1)):
                last_row -= 1
            last_row += 1

            # Lazy load a few utilities
            from pvevti.genutil import extractE9Serial, Date

            # Loop through the filedata list and add new excel rows
            for idx, row in filedata.iterrows():
                data = [extractE9Serial(row['File Name']), str(Date.extractDate(row['File Name'])),
                        row['Last Modified Time'], row['File Name'], row['File Size (KB)'], row['Route']]
                
                # If the row is occupied, add data
                if row.notna().any():
                    # Add each item to each appropriate cell
                    for col, item in enumerate(data):
                        sheet.cell(row=last_row + idx, column=col+1).value = item

                    # Mark as garbage in case something is under 1kb
                    if data[4] < 1:
                        sheet.cell(row=last_row + idx, column=14).value = "Yes"
                    
                    
            
            workbook.save(runlog)
            os.startfile(runlog)
        else:
            query = AllHasGoneAwry(self)
            if query.exec():
                self.configure_runlog(config_dir=False)
                self.execute_runlog()

    # Automatically pull in files from the last opened directory
    def autoscan(self):
        from pvevti.genutil import parseFileSize
        print("Clicked autoscan")
        num = 0
        print("Last Dir:",csv_open_default_path)
        if csv_open_default_path != "":
            files=[f for f in os.listdir(csv_open_default_path) if f.endswith('.csv')]
            all_items = [self.filesView.item(i).text() for i in range(self.filesView.count())]
            if all_items == ['Load CSV Files to get started.']:
                self.filesView.clear()
                global paths
                paths = []
            i = 0
            for file in files:
                self.updateProgress(i, len(files))
                i += 1
                sleep(0.01)
                name = csv_open_default_path+file
                if file not in all_items and name not in paths:
                    num += 1
                    sz = os.path.getsize(name)
                    self.filesView.addItem(QListWidgetItem("["+parseFileSize(sz)+"] "+name))
                    paths.append(name)
            print(paths)
            self.updateState('Imported '+str(num)+' file(s) from directory '+csv_open_default_path)
            self.updateProgress(i, len(files))
            QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)
    
    # Remove all the path items
    def clearItems(self):
        """
        Remove the loaded CSV names from the list.
        """
        # Only clear if the only item ISN'T 'Load CSV Files to get started.'
        all_items = [self.filesView.item(i).text() for i in range(self.filesView.count())]
        if all_items != ['Load CSV Files to get started.']:
            self.filesView.clear()
            self.filesView.addItem(QListWidgetItem("Load CSV Files to get started."))
            global paths
            paths = []
            self.updateState('Reset loaded CSVs.')
    
    # Remove only the selected path item
    def clearItem(self):
        """
        Remove the loaded CSV names from the list.
        """
        #all_items = [self.filesView.item(i).text() for i in range(self.filesView.count())]
        selected_items = self.filesView.selectedItems()
        if not selected_items:
            return
        global paths
        for item in selected_items:
            print(item.text())
            if item.text() != 'Load CSV Files to get started.':
                row = self.filesView.row(item)
                print(row)
                self.filesView.takeItem(row)
                self.updateState('Removed "'+paths[row]+'"')
                paths.pop(row)
        if len(paths) == 0:
            self.filesView.clear()
            self.filesView.addItem(QListWidgetItem("Load CSV Files to get started."))
            paths = []

    # Quick look feature (space)
    def quickView(self):
        print("QuickView")
        selected_items = self.filesView.selectedItems()
        if not selected_items:
            return
        global paths
        for item in selected_items:
            if item.text() != 'Load CSV Files to get started.':
                row = self.filesView.row(item)
                print(row,':',paths[row])
                popup = MapPopup(self, paths[row]).exec()

    ## Report generation
    # Create a metareport of the input files.
    def createMetaReport(self):
        from pvevti import genutil, csvutil
        print("METAREPORT")
        global paths
        data = [['Date', 'Serial', 'Run', 'Vehicle Data', 'Route']]
        sortedPaths = genutil.detectMatching(paths)
        self.updateProgress(1, 3)
        for path in sortedPaths:
            self.updateState("Create DF: "+path)
            df = csvutil.df_from_csv(path)
            data.append([str(genutil.Date.extractDate(path)), genutil.extractE9Serial(path), "GL"+genutil.extractGLRunNumber(path), "("+" ".join(genutil.extractVehicleData(genutil.extractE9Serial(path)))+")", str(genutil.getRoute(df, "Unknown")[0])])
        import csv
        self.updateProgress(2, 3)
        with open(savedir_rep+'/metareport.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)
        self.updateState("Saved to "+savedir_rep+'/metareport.csv')
        self.updateProgress(3, 3)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)
    
    # Create reports based on spec
    def makeReports(self):
        """
        Quick-develop reports based on the spec.
        """
        self.updateState("Import packages")
        from pvevti import pdfutil, csvutil
        from pvevti.genutil import extractE9Serial, extractGLRunNumber
        import json
        pdfutil.debugInfo = True

        self.updateState("Initialize")
        global paths
        global boxes
        global savedir_rep
        print("REPORTS: ", paths)
        print("Get default cfg")
        
        # Layout the configurations corresponding to the buttons
        self.updateState("Create default configurations")
        configs = [pdfutil.default_route_conditions_config, pdfutil.default_powersteering_pdf_config, pdfutil.default_cooling_pdf_config, 
                   pdfutil.default_config, pdfutil.default_electrical_pdf_config, pdfutil.default_hvac_pdf_config, pdfutil.default_towdyno_config]
        appends = ['Conditions', 'Powersteering', 'Cooling', 'Default', 'Electrical', 'HVAC', 'Dyno']
        self.updateState("Create custom configurations")
        for path in custom_reports:
            with open(path, 'r') as file:
                dictobj = json.load(file)
            configs.append(dictobj)
            appends.append(dictobj['docTitle'])
            print("   > Added {} to configs".format(dictobj['docTitle']))

        # Create a mask for configurations and appendages
        mask = []
        for box in boxes:
            mask.append(box.isChecked())

        # Develop said lists
        configs = [value for value, flag in zip(configs, mask) if flag]
        appends = [value for value, flag in zip(appends, mask) if flag]
        possibleReports = (1+len(configs)) * len(paths)
        i = 1
        # Loop through each item in the CSV list, and perform the requisite report creations.
        for path in paths:
            self.updateState("Create dataframe for "+path)
            df = csvutil.df_from_csv(path)
            self.updateProgress(i, possibleReports)
            i += 1
            for config, append in zip(configs, appends):
                name = (path.split('.csv')[0]+"_Report_" + append).split('/')[-1]
                self.updateState("Create "+append+" for "+name)
                print(name, 'in', rep_save_default_path)
                config['docTitle'] = name
                subtitle = extractE9Serial(path)+": GL"+extractGLRunNumber(path)+" ("+append+")"
                config['docSubTitle'] = subtitle
                pdfutil.createDocument(df, config, rep_save_default_path)
                self.updateProgress(i, possibleReports)
                i += 1
        savedir_rep = rep_save_default_path
        self.updateState(str(len(paths) * len(configs))+" reports complete.")
        self.savedRep.setVisible(True)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)
    

    ## Processing
    # Performs a merge on a list of files.
    # If PQR files are present, merges them into a sorted, merged CSV.
    # If SLC files are present, creates a single SLC with all issues, grouped by vehicle and sorted by run number.
    # Both can occur simultaneously.
    def mergeFiles(self, toSort=[], toName=""):
        modifiers = QApplication.queryKeyboardModifiers()
        if modifiers & Qt.ShiftModifier:
            print("Pressed with shift held.")
            MergePreferences(self).exec()
            

        self.updateState("Import packages")
        from pvevti import genutil, csvutil
        from csv import writer
        self.updateState("Initialize")
        if toSort == [] or toSort == False:
            global paths
        else:
            paths = toSort
        global savedir_csv
        print("toSort",toSort,'\npaths',paths)
        # Define two classifications of mergeable CSVs
        paths_SLC = [path for path in paths if "SLC" in path]
        paths_PQR = [path for path in paths if "PQR" in path]

        # Reservations
        dfs = []
        gls = []
        pathname = ""

        # Fill Dataframe list and Run number list
        self.updateState("Sort files")
        for idx, path in enumerate(genutil.detectMatching(paths_PQR)):
            try:
                if idx == 0:
                    pathname = path
                dfs.append(csvutil.df_from_csv(path, filterSignals))
                gls.append(genutil.extractGLRunNumber(path))
                self.updateState("Added "+str(genutil.extractGLRunNumber(path)))
                print(path)
            except:
                print("Failed to cast {} into a dataFrame.".format(path))

        # Combine PQR CSV files
        self.updateState("Combine CSVs")
        self.updateProgress(1, 3)
        df = genutil.combineDFs(dfs, gls)
        self.updateProgress(2, 3)
        if len(df) > 0:

            # Extract name and update state
            name = pathname.split('/')[-1]
            self.updateState("Saving to "+str(csv_save_default_path+'/'+name.split('.csv')[0]+"_Combined.csv"))
            print("PQR SAVE",csv_save_default_path)

            # Save
            csvutil.df_to_csv(df, csv_save_default_path+'/'+name, addition="_Combined")
            self.updateProgress(3, 3)
            savedir_csv = csv_save_default_path
            self.updateState("Save to "+csv_save_default_path+'/'+name.split('.csv')[0]+"_Combined.csv "+" complete.")
            self.savedCSV.setVisible(True)
        else:
            print("NO PQR SAVE")
            self.updateProgress(3, 3)
            self.updateState("No PQR files provided. No merge output.")
        
        # Combine SLC CSV Files
        if len(paths_SLC) > 0:

            uniqueList = []

            # Define CSV architecture
            maxIter = len(paths_SLC)+1
            self.updateProgress(1, maxIter)
            data = [["Vehicle", "Run Number", "Date", "Signal", "Recorded Value (Limit)"]]
            vehicles = list(set([genutil.extractE9Serial(path) for path in paths_SLC]))
            print("UNIQUE VEHICLES "+str(vehicles))

            # Group by vehicle, sort by run number. 
            i = 2
            for veh in vehicles:
                paths_SLC_veh = [path for path in paths_SLC if veh in path]
                runs = [int(genutil.extractGLRunNumber(path)) for path in paths_SLC_veh]
                paths_SLC_veh = [path for run, path in sorted(zip(runs, paths_SLC_veh))]
                for path in paths_SLC_veh:
                    self.updateProgress(i, maxIter)
                    # Define important properties and add list elements to reflect
                    try:
                        items =     csvutil.df_from_csv(path).values.tolist()
                        vehicle =   genutil.extractE9Serial(path)
                        runNumber = genutil.extractGLRunNumber(path)
                        date =      genutil.Date.extractDate(path)
                        i += 1
                        for item in items:
                            data.append([vehicle, runNumber, str(date), item[0], str(item[4])+" "+item[1]+" (Limit: "+str(item[3])+")"])
                            strAdd = vehicle+": "+item[0]
                            if strAdd not in uniqueList:
                                uniqueList.append(strAdd)
                    except:
                        print("Failed to extract information from {}.".format(path))
            
            # Save to a new CSV
            if toName == "":
                toName = csv_save_default_path+"SLC_Combined.csv"
            elif ".csv" not in toName:
                toName = csv_save_default_path + toName + ".csv"
            else:
                toName = csv_save_default_path + toName
            with open(toName, 'w', newline='') as file:
                w = writer(file)
                w.writerows(data)
            self.updateState("Saved to "+csv_save_default_path+"SLC_Combined.csv")
            savedir_csv = csv_save_default_path
            self.savedCSV.setVisible(True)
            if len(uniqueList) > 0:
                openFile = InfoPopup(self, "SLC Results", "Out-of-Bounds signals from the specified day:", uniqueList)
                if openFile.exec() == 1:
                    os.startfile(toName)
            else:
                from PySide6.QtWidgets import QMessageBox
                msg_box = QMessageBox(self)
                msg_box.setIcon(QMessageBox.Information)  # Set the icon to Information
                msg_box.setWindowTitle("No out-of-bounds values!")
                msg_box.setText("There were no out-of-bounds values detected in the provided {} SLC file(s).".format(len(paths_SLC)))
                msg_box.setInformativeText("You can add more details here if needed.")
                msg_box.setStandardButtons(QMessageBox.Ok) # Add an OK button
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)

    # Create CSV files based on the spec
    def processFiles(self):
        """
        Perform CSV operations based on the spec.
        """
        self.updateState("Import packages")
        from pvevti import genutil, csvutil, jsonutil

        # Get preferences and initialize
        prefs = jsonutil.Prefs.getPrefs()
        self.updateState("Initialize")
        global paths
        global savedir_csv
        maxIter = len(paths) * 4 + 1
        i = 1
        print("FILES: ", paths)
        
        # Loop through each provided path
        for path in paths:
            self.updateProgress(i, maxIter)
            i += 1
            self.updateState("Get metadata")
            print("\n PATH PROCESS:",path)
            name = path.split('/')[-1]
            print(" == DATA ==")
            df = csvutil.df_from_csv(path)

            # Extract metadata
            print("  Metadata")
            metadata = {
                "Serial": genutil.extractE9Serial(path),
                "RunNumber": genutil.extractGLRunNumber(path),
                "VehicleData": genutil.extractVehicleData(genutil.extractE9Serial(path))
            }
            print("  ",metadata)
            print("  Route\n   "+str(genutil.getRoute(df)))
            print(" == PROCESSING ==")
            rounding_accuracy = jsonutil.Prefs.getRoundingAcc(prefs, df.columns)
            self.updateProgress(i, maxIter)
            i += 1
            if self.compression.currentText() != 'Disabled':
                self.updateState("{} (GL{}): Compression".format(metadata['Serial'], metadata['RunNumber']))
                print("ROUND")
                df = genutil.roundCols(df, rounding_accuracy)
                if self.compression.currentText() in ['Normal', 'Aggressive']:
                    self.updateState("{} (GL{}): Compression (drop)".format(metadata['Serial'], metadata['RunNumber']))
                    print("DROP")
                    df = genutil.discard(df, prefs)
                self.updateProgress(i, maxIter)
                i += 1
            else:
                i+=1
            
            if self.gps.currentText() == 'Enabled':
                self.updateState("{} (GL{}): GPS Filter".format(metadata['Serial'], metadata['RunNumber']))
                print('GPS FILTER')
                df = genutil.gps_filter_data(df)
                self.updateProgress(i, maxIter)
                i += 1
            else:
                i+=1
            
            if doExport:
                self.updateState("{} (GL{}): Export".format(metadata['Serial'], metadata['RunNumber']))
                if replaceOriginals:
                    print("SAVE replace original at",path)
                    csvutil.df_to_csv(df, path, addition="")
                else:
                    print("SAVE to new directory",csv_save_default_path+'/'+name)
                    csvutil.df_to_csv(df, csv_save_default_path+'/'+name)
                self.updateProgress(i, maxIter)
                i += 1
            else:
                i+=1
        savedir_csv = csv_save_default_path
        self.updateState(str(len(paths))+" csv file(s) processed.")
        self.savedCSV.setVisible(True)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)

    # Create new CSV files and develop reports based on the resultant data 
    def processAndReport(self):
        """
        Quick-develop reports based on the spec.
        """
        self.updateState("Import packages")
        from pvevti import genutil, csvutil, jsonutil, pdfutil
        from pvevti.genutil import extractE9Serial, extractGLRunNumber
        import json

        self.updateState("Initialize")
        global paths, boxes
        global savedir_csv, savedir_rep
        print("REPORTS: ", paths)
        print("Get default cfg")
        
        # Layout the configurations corresponding to the buttons
        self.updateState("Create default configurations")
        configs = [pdfutil.default_route_conditions_config, pdfutil.default_powersteering_pdf_config, pdfutil.default_cooling_pdf_config, 
                   pdfutil.default_config, pdfutil.default_electrical_pdf_config, pdfutil.default_hvac_pdf_config, pdfutil.default_towdyno_config]
        appends = ['Conditions', 'Powersteering', 'Cooling', 'Default', 'Electrical', 'HVAC', 'Dyno']
        self.updateState("Create custom configurations")
        for path in custom_reports:
            with open(path, 'r') as file:
                dictobj = json.load(file)
            configs.append(dictobj)
            appends.append(dictobj['docTitle'])
            print("   > Added {} to configs".format(dictobj['docTitle']))

        # Create a mask for configurations and appendages
        mask = []
        for box in boxes:
            mask.append(box.isChecked())

        # Develop said lists
        configs = [value for value, flag in zip(configs, mask) if flag]
        appends = [value for value, flag in zip(appends, mask) if flag]
        maxIter = len(paths) * 4 + len(configs) * len(paths) + 1
        i = 1
        prefs = jsonutil.Prefs.getPrefs()

        print("FILES: ", paths)
        
        for path in paths:
            self.updateProgress(i, maxIter)
            i += 1
            self.updateState("Create dataframe for "+path)
            print("\n PATH PROCESS:",path)
            name = path.split('/')[-1]
            print(" == DATA ==")
            df = csvutil.df_from_csv(path)
            self.updateState("Get metadata")
            print("  Metadata")
            metadata = {
                "Serial": genutil.extractE9Serial(path),
                "RunNumber": genutil.extractGLRunNumber(path),
                "VehicleData": genutil.extractVehicleData(genutil.extractE9Serial(path))
            }
            print("  ",metadata)
            print("  Route\n   "+str(genutil.getRoute(df)))
            print(" == PROCESSING ==")
            rounding_accuracy = jsonutil.Prefs.getRoundingAcc(prefs, df.columns)
            if self.compression.currentText() != 'Disabled':
                self.updateState("{} (GL{}): Compression".format(metadata['Serial'], metadata['RunNumber']))
                print("ROUND")
                df = genutil.roundCols(df, rounding_accuracy)
                if self.compression.currentText() in ['Normal', 'Aggressive']:
                    self.updateState("{} (GL{}): Compression (drop)".format(metadata['Serial'], metadata['RunNumber']))
                    print("DROP")
                    df = genutil.discard(df, prefs)
                self.updateProgress(i, maxIter)
                i += 1
            else:
                i += 1
            
            if self.gps.currentText() == 'Enabled':
                self.updateState("{} (GL{}): GPS Filter".format(metadata['Serial'], metadata['RunNumber']))
                print('GPS FILTER')
                df = genutil.gps_filter_data(df)
                self.updateProgress(i, maxIter)
                i += 1
            else:
                i += 1

            if doExport:
                self.updateState("{} (GL{}): Export".format(metadata['Serial'], metadata['RunNumber']))
                if replaceOriginals:
                    print("SAVE replace original at",path)
                    csvutil.df_to_csv(df, path, addition="")
                else:
                    print("SAVE to new directory",csv_save_default_path+'/'+name)
                    csvutil.df_to_csv(df, csv_save_default_path+'/'+name)
                self.updateProgress(i, maxIter)
                i += 1
            else:
                i += 1
            
            for config, append in zip(configs, appends):
                print("REPORT",append)
                name = (path.split('.csv')[0]+"_Report_" + append).split('/')[-1]
                self.updateState("Create "+append+" for "+name)
                print(name, 'in', rep_save_default_path)
                config['docTitle'] = name
                subtitle = extractE9Serial(path)+": GL"+extractGLRunNumber(path)+" ("+append+")"
                config['docSubTitle'] = subtitle
                pdfutil.createDocument(df, config, rep_save_default_path)
                self.updateProgress(i, maxIter)
                i += 1
        savedir_rep = rep_save_default_path
        if replaceOriginals:
            savedir_csv = '/'.join(paths[-1].split('/')[0:-1])
        else:
            savedir_csv = csv_save_default_path
        self.updateState(str(len(paths))+" csv file(s) processed. "+str(len(configs)*len(paths))+" report(s) generated.")
        self.savedCSV.setVisible(True)
        self.savedRep.setVisible(True)
        self.updateProgress(i, maxIter)
        QApplication.setOverrideCursor(Qt.CursorShape.ArrowCursor)

    # One can infer
    def closeApp(self):
        """
        Close the application.
        """
        print('Closing FPT.')
        self.close()

    # One can also infer
    def restartApp(self):
        import subprocess
        print('Restarting!')
        QApplication.quit()
        subprocess.Popen([sys.executable, sys.argv[0]])

# Profiling timer class, unused
class Timer():
    def __init__(self):
        import time
        self.startTime = time.time()
        self.currentTime = time.time()
    
    def elapsed(self, label=""):
        import time
        label = " [{}]".format(label) if label != "" else ""
        print("Elapsed{}: {:.2f}".format(label, time.time()-self.currentTime))
        self.currentTime = time.time()
    
    def total(self):
        import time
        print("Total Elapsed: {:.2f}".format(time.time()-self.startTime))

# Main
if __name__ == '__main__':
    # Profiling (unused)
     # timer = Timer()

    # Create application, set window icon
    app = QtWidgets.QApplication(sys.argv)
     # timer.elapsed("App create")
    app.setWindowIcon(QtGui.QIcon(os.path.join(basedir, 'icon_dev.ico')))
     # timer.elapsed("Window Icon")

    # Clear out the scrungy looking tooltips
    app.setStyleSheet('QToolTip {color:black;background-color:white;border:1px solid #D0D0D0;border-radius:0px;font-size:10px;}')
     # timer.elapsed("Stylesheet")

    # Create and render a temporary splash screen while resources load
    mp = QPixmap(resource_path('icons/logo.png'))
     # timer.elapsed("Splash Image Load")
    splash = QSplashScreen(mp)
     # timer.elapsed("Splash Create")
    splash.show()
     # timer.elapsed("Splash Show")
    app.processEvents()
     # timer.elapsed("App Update")
    load_images(splash, app)
     # timer.elapsed("Load Images")
     # timer.total()

    # Create the main window and close the splash screen once it's open
    w = MainWindow()
    splash.finish(w)

    # Execute
    app.exec()